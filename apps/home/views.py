# -*- encoding: utf-8 -*-
"""
Copyright (c) 2019 - present AppSeed.us
"""
from datetime import datetime, timedelta
from django import template
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse, HttpResponseForbidden
from django.shortcuts import render
from django.template import loader
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from apps.authentication.models import CustomUser, Role
from django.db import models
from decouple import config

from .models import  VoucherType, Lots, DiningRoom, ClientDiner, Client, Employee, PayrollType, Entry, EmployeeClientDiner, Voucher
from apps.authentication.models import CustomUser, Role
from django.db.models import Q, F, Count, Exists, OuterRef
from django.db import IntegrityError, transaction
import json
import re
import pytz
from django.contrib.auth.hashers import make_password
from django.shortcuts import get_object_or_404
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from .transactions.clients import change_client_status
from .transactions.comedores import change_dining_room_status
import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
from .admin import admin_views, user_views
import os
from email.message import EmailMessage
import ssl
import smtplib
from apps.pdf_generation import generate_qrs_pdf, prepare_qrs, generate_lot_pdf, prepare_url_pdf, clean_pdf_dir, prepare_qr, generate_perpetual_voucher_pdf, verify_lot_pdf_exists, verify_voucher_pdf_exists, create_lot_pdf_name, create_voucher_pdf_name
import re

@login_required(login_url="/login/")
def index(request):
    context = {'segment': 'index'}

    html_template = loader.get_template('home/index.html')
    return HttpResponse(html_template.render(context, request))


@login_required(login_url="/login/")
def pages(request):
    
    context = {'segment': request.path.split('/')[-1]}
    # All resource paths end in .html.
    # Pick out the html file name from the url. And load that template.
    try:

        load_template = request.path.split('/')[-1]

        if load_template in map(lambda x: x + '.html', admin_views) and request.user.role_id != 1:
           response = render(request, 'home/page-403.html') 
           return HttpResponseForbidden(response.content)

        if load_template in map(lambda x: x + '.html', user_views) and request.user.role_id != 2:
            response = render(request, 'home/page-403.html')
            return HttpResponseForbidden(response.content)

        if 'reporte' in load_template and request.user.role_id == 1:
            html_template = loader.get_template('home/reportes/' + load_template)
        else:
            html_template = loader.get_template('home/' + load_template)
        
        return HttpResponse(html_template.render(context, request))

    except template.TemplateDoesNotExist:
        html_template = loader.get_template('home/page-404.html')
        return HttpResponse(html_template.render(context, request))

    except:
        html_template = loader.get_template('home/page-500.html')
        return HttpResponse(html_template.render(context, request))

# ============================= COMEDORES =============================
@csrf_exempt
def get_all_comedores(request):
    try:
        # Obtener el valor del filtro de la solicitud
        client_value = request.GET.get('client-id', 'all')

        # Obtener todos los comedores con la información del cliente y del encargado
        dining_rooms_query = ClientDiner.objects.select_related('client', 'dining_room', 'dining_room__in_charge').values(
            'client__company',
            'dining_room__id',
            'dining_room__name',
            'dining_room__description',
            'dining_room__in_charge__first_name',
            'dining_room__in_charge__last_name',
            'dining_room__status'
        ).distinct()

        # Aplicar el filtro si no es 'all'
        if client_value != 'all':
            dining_rooms_query = dining_rooms_query.filter(client__id=client_value)

        # Renombrar campos para evitar confusión
        dining_rooms_list = [
            {
                'company': dr['client__company'],
                'id': dr['dining_room__id'],
                'name': dr['dining_room__name'],
                'description': dr['dining_room__description'],
                'in_charge_first_name': dr['dining_room__in_charge__first_name'],
                'in_charge_last_name': dr['dining_room__in_charge__last_name'],
                'status': dr['dining_room__status']
            }
            for dr in dining_rooms_query
        ]


        return JsonResponse({"comedores": dining_rooms_list})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    

@csrf_exempt
def get_comedores(request):
    try:
        # Obtener el valor del filtro de la solicitud
        filter_value = request.GET.get('filter', 'all')

        # Obtener todos los comedores con la información del cliente y del encargado
        dining_rooms_query = DiningRoom.objects.select_related('in_charge').prefetch_related(
            'client_diner_dining_room__client'
        ).values(
            'id',
            'name',
            'description',
            'status',
            'in_charge__id',
            'in_charge__first_name',
            'in_charge__last_name',
            'client_diner_dining_room__client__company'
        ).distinct()

        # Aplicar el filtro si no es 'all'
        if filter_value != 'all':
            dining_rooms_query = dining_rooms_query.filter(client_diner_dining_room__client__id=filter_value)

        # Renombrar campos para evitar confusión
        dining_rooms_list = [
            {
                'id': dr['id'],
                'name': dr['name'],
                'description': dr['description'],
                'status': dr['status'],
                'in_charge_id': dr['in_charge__id'],
                'in_charge_first_name': dr['in_charge__first_name'],
                'in_charge_last_name': dr['in_charge__last_name'],
                'company': dr['client_diner_dining_room__client__company']
            }
            for dr in dining_rooms_query
        ]

        # Obtener lista de clientes únicos
        clients = Client.objects.values('id', 'company').distinct()

        # Paginación
        page_number = request.GET.get('page', 1)
        paginator = Paginator(dining_rooms_list, 10)  # 10 comedores por página
        page_obj = paginator.get_page(page_number)

        context = {
            'dining_rooms': list(page_obj),
            'clients': list(clients),
            'page_number': page_obj.number,
            'num_pages': paginator.num_pages,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous(),
        }

        return JsonResponse(context)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def get_comedor(request):
    try:
        dining_room_id = request.GET.get('dining_room_id')

        # Validar que se proporcione el ID del comedor
        if not dining_room_id:
            return JsonResponse({'error': 'El ID del comedor es obligatorio'}, status=400)

        # Realizar la consulta con las uniones necesarias
        dining_room = DiningRoom.objects.filter(id=dining_room_id).select_related('in_charge').prefetch_related(
            'client_diner_dining_room__client'
        ).values(
            'id',
            'name',
            'description',
            'status',
            'in_charge__id',
            'in_charge__first_name',
            'in_charge__last_name',
            'client_diner_dining_room__client__id',
            'client_diner_dining_room__client__company',
            'client_diner_dining_room__client__status'
        ).first()

        # Verificar si el comedor existe
        if not dining_room:
            return JsonResponse({'error': 'Comedor no encontrado'}, status=404)

        # Formatear los datos del encargado
        in_charge = {
            'id': dining_room['in_charge__id'],
            'first_name': dining_room['in_charge__first_name'],
            'last_name': dining_room['in_charge__last_name']
        } if dining_room['in_charge__id'] else None

        # Formatear los datos del cliente
        client = {
            'id': dining_room['client_diner_dining_room__client__id'],
            'company': dining_room['client_diner_dining_room__client__company'],
            'status': dining_room['client_diner_dining_room__client__status']
        }

        # Construir el contexto de respuesta
        context = {
            'dining_room_id': dining_room['id'],
            'name': dining_room['name'],
            'description': dining_room['description'],
            'status': dining_room['status'],
            'in_charge': in_charge,
            'client': client
        }

        return JsonResponse(context)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def create_comedor(request):
    try:
        data = json.loads(request.body)
        name = data.get('name')
        description = data.get('description')
        status = data.get('status')
        in_charge = data.get('in_charge')
        client_id = data.get('client')  # Obtener el client_id del request
        created_by_id = request.user.id

        # Validar que el nombre no esté vacío y tenga una longitud válida
        if not name or len(name) < 3 or len(name) > 50:
            return JsonResponse({'message': 'El nombre debe tener entre 3 y 50 caracteres', 'status': 'danger'}, status=400)

        # Validar que la descripción no esté vacía y tenga una longitud válida
        if not description or len(description) > 100:
            return JsonResponse({'message': 'La descripción no puede tener más de 100 caracteres', 'status': 'danger'}, status=400)

        # Validar que el cliente exista y esté activo
        client = Client.objects.filter(id=client_id, status=True).first()
        if not client:
            return JsonResponse({'message': 'El cliente proporcionado no existe o está inactivo', 'status': 'danger'}, status=400)

        # Validar que el encargado exista y esté activo (si se proporciona)
        if in_charge:
            encargado = CustomUser.objects.filter(id=in_charge, status=True).first()
            if not encargado:
                return JsonResponse({'message': 'El encargado proporcionado no existe o está inactivo', 'status': 'danger'}, status=400)

        # Crear el comedor en el modelo DiningRoom
        dining_room = DiningRoom.objects.create(
            name=name,
            description=description,
            status=status,
            in_charge_id=in_charge if in_charge else None,
            created_by_id=created_by_id,
            updated_by_id=created_by_id
        )

        # Crear la entrada en el modelo ClientDiner
        ClientDiner.objects.create(
            dining_room_id=dining_room.id,
            client_id=client_id,
            created_by_id=created_by_id,
            updated_by_id=created_by_id
        )

        return JsonResponse({'message': 'Comedor creado exitosamente', 'status': 'success'}, status=201)
    except Exception as e:
        return JsonResponse({'message': str(e), 'status': 'danger'}, status=500)


@csrf_exempt
def update_comedor(request):
    try:
        user_id = request.user.id
        data = json.loads(request.body)
        dining_room_id = data.get('dining_room_id')
        name = data.get('name')
        description = data.get('description')
        status = data.get('status')
        in_charge = data.get('inCharge')
        client_id = data.get('client')  # Obtener el client_id del request

        # Validar que el comedor exista
        dining_room = DiningRoom.objects.filter(id=dining_room_id).first()
        if not dining_room:
            return JsonResponse({'message': 'Comedor no encontrado', 'status': 'danger'}, status=404)

        # Validar que el nombre no esté vacío y tenga una longitud válida
        if not name or len(name) < 3 or len(name) > 50:
            return JsonResponse({'message': 'El nombre debe tener entre 3 y 50 caracteres', 'status': 'danger'}, status=400)

        # Validar que la descripción no esté vacía y tenga una longitud válida
        if not description or len(description) > 100:
            return JsonResponse({'message': 'La descripción no puede tener más de 100 caracteres', 'status': 'danger'}, status=400)

        # Validar que el cliente exista y esté activo
        client = Client.objects.filter(id=client_id, status=True).first()
        if not client:
            return JsonResponse({'message': 'El cliente proporcionado no existe o está inactivo', 'status': 'danger'}, status=400)

        # Validar que el encargado exista y esté activo (si se proporciona)
        if in_charge:
            encargado = CustomUser.objects.filter(id=in_charge, status=True).first()
            if not encargado:
                return JsonResponse({'message': 'El encargado proporcionado no existe o está inactivo', 'status': 'danger'}, status=400)

        with transaction.atomic():
            # Actualizar los campos del comedor
            dining_room.name = name
            dining_room.description = description
            dining_room.status = status

            # Si el comedor se actualiza a inactivo, eliminar el encargado
            if not status:  # status es False
                dining_room.in_charge = None
            else:
                # Permitir que in_charge sea nulo si no está inactivo
                dining_room.in_charge_id = in_charge if in_charge else None

            dining_room.updated_by_id = user_id
            dining_room.save()

            # Actualizar o crear la entrada en el modelo ClientDiner
            ClientDiner.objects.update_or_create(
                dining_room_id=dining_room_id,
                defaults={
                    'client_id': client_id,
                    'updated_by_id': user_id
                }
            )

            # Si el comedor se desactiva, desactivar también a los empleados asignados
            if not status:
                client_diners = ClientDiner.objects.filter(dining_room=dining_room).all()
                for client_diner in client_diners:
                    client_diner.status = False
                    client_diner.updated_by_id = user_id
                    client_diner.save()

                    employee_client_diners = EmployeeClientDiner.objects.filter(client_diner=client_diner).all()
                    for employee_client_diner in employee_client_diners:
                        employee_client_diner.status = False
                        employee_client_diner.updated_by_id = user_id
                        employee_client_diner.save()

                        employee = employee_client_diner.employee
                        employee.status = False
                        employee.updated_by_id = user_id
                        employee.save()

        return JsonResponse({'message': 'Comedor actualizado correctamente', 'status': 'success'})
    except Exception as e:
        return JsonResponse({'message': str(e), 'status': 'danger'}, status=500)

@csrf_exempt
def get_encargados(request):
    try:
        # Filtrar los usuarios con role_id = 2
        encargados = CustomUser.objects.filter(role_id=2, status=True)
        
        # Excluir los usuarios que están asignados como in_charge en cualquier DiningRoom
        encargados = encargados.exclude(dining_room_in_charge__isnull=False)
        
        # Seleccionar los campos necesarios
        encargados = encargados.values('id', 'first_name', 'last_name')
        
        context = {
            'encargados': list(encargados)
        }

        return JsonResponse(context)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    

# Obtiente los clientes que tienen algun comedor asignado (select del filtro)   
def get_clientes_comedores(request):
    try:
        # Obtener los clientes que tienen comedores asignados
        clientes = ClientDiner.objects.select_related('client').values('client__id', 'client__company').distinct()
        
        # Formatear los datos para la respuesta JSON
        clientes_list = [
            {
                'id': cliente['client__id'],
                'company': cliente['client__company']
            }
            for cliente in clientes
        ]
        
        context = {
            'clientes': clientes_list
        }
        
        return JsonResponse(context)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
        
# ================================== Clientes ================================== #
@csrf_exempt
def client_list(request):
    if request.method == 'GET':
        search_query = request.GET.get('search', '')
        clients = Client.objects.all().order_by('id').values('id', 'company', 'name', 'lastname', 'second_lastname', 'rfc', 'email', 'phone', 'address', 'status')

        if search_query:
            clients = clients.filter(
                Q(company__icontains=search_query) |
                Q(name__icontains=search_query) |
                Q(lastname__icontains=search_query) |
                Q(second_lastname__icontains=search_query)
            )

        page = request.GET.get('page', 1)
        paginator = Paginator(clients, 10)  # 10 clientes por página

        try:
            clients_page = paginator.page(page)
        except PageNotAnInteger:
            clients_page = paginator.page(1)
        except EmptyPage:
            clients_page = paginator.page(paginator.num_pages)

        return JsonResponse({
            'clients': list(clients_page),
            'page': clients_page.number,
            'pages': paginator.num_pages
        })
    elif request.method == 'POST':
        try:
            id_user = request.user.id
            if not CustomUser.objects.filter(id=id_user, role__name='Administrador').exists():
                return JsonResponse({'error': 'No tienes permiso para crear un cliente'}, status=403)
            data = json.loads(request.body)

            if len(data['company']) < 2:
                return JsonResponse({'error': 'El nombre de la compañía debe tener al menos 2 caracteres'}, status=400)
            
            if len(data['company']) > 50:
                return JsonResponse({'error': 'El nombre de la compañía debe tener máximo 50 caracteres'}, status=400)

            if len(data['name']) < 2 or len(data['lastname']) < 2:
                return JsonResponse({'error': 'El nombre y apellido paterno deben tener al menos 2 caracteres'}, status=400)
            
            if len(data['rfc']) != 13:
                return JsonResponse({'error': 'El RFC debe tener 13 caracteres'}, status=400)
            
            try:
                int(data['phone'])
            except ValueError:
                if not data['phone'].isdigit():
                    return JsonResponse({'error': 'El teléfono debe contener solo dígitos'}, status=400)
            
            if len(data['phone']) != 10:
                return JsonResponse({'error': 'El teléfono debe tener 10 dígitos'}, status=400)
            
            if len(data['address']) < 5:
                return JsonResponse({'error': 'La dirección debe tener al menos 5 caracteres'}, status=400)
            
            if len(data['address']) > 100:
                return JsonResponse({'error': 'La dirección debe tener máximo 100 caracteres'}, status=400)
                        
            if '@' not in data['email'] or '.' not in data['email']:
                return JsonResponse({'error': 'Correo electrónico inválido'}, status=400)
            
            client = Client.objects.create(
                company=data['company'],
                name=data['name'],
                lastname=data['lastname'],
                second_lastname=data.get('second_lastname', ''),
                rfc=data['rfc'].upper(),
                email=data['email'],
                phone=data['phone'],
                address=data['address'],
                status=data.get('status', True),
                created_by=request.user,
                updated_by_id=request.user.id
            )
            return JsonResponse({'message': 'Cliente creado correctamente', 'client_id': client.id}, status=201)
        except IntegrityError as e:
            if 'email' in str(e):
                return JsonResponse({'error': 'El correo electrónico ya existe'}, status=400)
            else:
                return JsonResponse({'error': 'Error de integridad de datos'}, status=400)
        except KeyError as e:
            return JsonResponse({'error': f'Campo faltante: {str(e)}'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

@csrf_exempt
def client_detail(request, client_id):
    client = get_object_or_404(Client, id=client_id)
    if request.method == 'GET':
        try:
            client_data = Client.objects.filter(id=client_id).values('id', 'company', 'name', 'lastname', 'second_lastname', 'rfc', 'email', 'phone', 'address', 'status', 'created_by', 'updated_by').first()
            return JsonResponse(client_data)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            if len(data['company']) < 2:
                return JsonResponse({'error': 'El nombre de la compañía debe tener al menos 2 caracteres'}, status=400)
            
            if len(data['company']) > 50:
                return JsonResponse({'error': 'El nombre de la compañía debe tener máximo 50 caracteres'}, status=400)

            if len(data['name']) < 2 or len(data['lastname']) < 2:
                return JsonResponse({'error': 'El nombre y apellido paterno deben tener al menos 2 caracteres'}, status=400)
            
            if len(data['rfc']) != 13:
                return JsonResponse({'error': 'El RFC debe tener 13 caracteres'}, status=400)
            
            if len(data['phone']) != 10:
                return JsonResponse({'error': 'El teléfono debe tener 10 caracteres'}, status=400)
            
            if len(data['address']) < 5:
                return JsonResponse({'error': 'La dirección debe tener al menos 5 caracteres'}, status=400)
            
            if len(data['address']) > 100:
                return JsonResponse({'error': 'La dirección debe tener máximo 100 caracteres'}, status=400)
                        
            if '@' not in data['email'] or '.' not in data['email']:
                return JsonResponse({'error': 'Correo electrónico inválido'}, status=400)
            
            client.company = data.get('company', client.company)
            client.name = data.get('name', client.name)
            client.lastname = data.get('lastname', client.lastname)
            client.second_lastname = data.get('second_lastname', client.second_lastname)
            client.rfc = data.get('rfc', client.rfc).upper()
            client.email = data.get('email', client.email)
            client.phone = data.get('phone', client.phone)
            client.address = data.get('address', client.address)
            
            user_who_updated = CustomUser.objects.filter(id=request.user.id).first()
            if user_who_updated is None:
                return JsonResponse({'error': 'Usuario no encontrado'})
            
            if client.status != data['status']:
                change_client_status(user_who_updated, client, data['status'])
            
            client.updated_by = user_who_updated
            client.save()
            return JsonResponse({'message': 'Cliente actualizado correctamente'})
        except IntegrityError as e:
            if 'email' in str(e):
                return JsonResponse({'error': 'El correo electrónico ya existe'}, status=400)
            else:
                return JsonResponse({'error': 'Error de integridad de datos'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
        
# ================================== Usuarios ================================== #
@csrf_exempt
def user_list(request):
    if request.method == 'GET':
        search_query = request.GET.get('search', '')
        role_filter = request.GET.get('role', '')

        users = CustomUser.objects.all().order_by('id').values(
            'id', 'username', 'first_name', 'last_name', 'second_last_name', 'email', 'role__name', 'dining_room_in_charge__name', 'dining_room_in_charge__client_diner_dining_room__client__company', 'status'
        ).distinct()

        if search_query:
            users = users.filter(
                Q(username__icontains=search_query) |
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query) |
                Q(second_last_name__icontains=search_query)
            )

        if role_filter:
            users = users.filter(role__id=role_filter)

        page = request.GET.get('page', 1)
        paginator = Paginator(users, 10)  # 10 usuarios por página

        try:
            users_page = paginator.page(page)
        except PageNotAnInteger:
            users_page = paginator.page(1)
        except EmptyPage:
            users_page = paginator.page(paginator.num_pages)

        return JsonResponse({
            'users': list(users_page),
            'page': users_page.number,
            'pages': paginator.num_pages
        })
    elif request.method == 'POST':
        try:
            id_user = request.user.id
            if not CustomUser.objects.filter(id=id_user, role__name='Administrador').exists():
                return JsonResponse({'error': 'No tienes permiso para crear un usuario'}, status=403)
            data = json.loads(request.body)
            role = get_object_or_404(Role, id=data['role_id'])

            if len(data['username']) < 5:
                return JsonResponse({'error': 'El nombre de usuario debe tener al menos 5 caracteres'}, status=400)

            if len(data['first_name']) < 2 or len(data['last_name']) < 2:
                return JsonResponse({'error': 'El nombre y apellido paterno deben tener al menos 2 caracteres'}, status=400)
                        
            if 'email' in data and data['email']:
                if '@' not in data['email'] or '.' not in data['email']:
                    return JsonResponse({'error': 'Correo electrónico inválido'}, status=400)
            
            if len(data['password']) < 8 or not any(char.isdigit() for char in data['password']) or not any(char.isalpha() for char in data['password']):
                return JsonResponse({'error': 'La contraseña debe tener al menos 8 caracteres, una letra y un número'}, status=400)

            with transaction.atomic():
                user = CustomUser.objects.create(
                    username=data['username'],
                    first_name=data['first_name'],
                    last_name=data['last_name'],
                    second_last_name=data['second_last_name'],
                    email=data.get('email', ''),
                    password=make_password(data['password']),
                    role=role,
                    status=data.get('status', True),
                    created_by=request.user
                )
                if 'dining_room_in_charge' in data and data['dining_room_in_charge'] is not None:
                    if data['dining_room_in_charge'] == 'null' or data['dining_room_in_charge'] == 'no':
                        if user.dining_room_in_charge.exists():
                            dining_room = user.dining_room_in_charge.first()
                            dining_room.in_charge = None
                            dining_room.save()
                        user.dining_room_in_charge.clear()
                    else:
                        dining_room = get_object_or_404(DiningRoom, id=data['dining_room_in_charge'])
                        dining_room.in_charge = user
                        dining_room.save()
                        user.dining_room_in_charge.set([dining_room])

                if data['status'] == False:
                    dining_rooms = DiningRoom.objects.filter(in_charge=user)
                    for dining_room in dining_rooms:
                        dining_room.in_charge = None
                        dining_room.save()
                        user.dining_room_in_charge.clear()
                user.save()

            return JsonResponse({'message': 'Usuario creado correctamente', 'user_id': user.id}, status=201)
        except IntegrityError as e:
            if 'username' in str(e):
                return JsonResponse({'error': 'El nombre de usuario ya existe'}, status=400)
            elif 'email' in str(e):
                return JsonResponse({'error': 'El correo electrónico ya existe'}, status=400)
            else:
                return JsonResponse({'error': 'Error de integridad de datos'}, status=400)
        except KeyError as e:
            return JsonResponse({'error': f'Campo faltante: {str(e)}'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

@csrf_exempt
def user_detail(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)
    if request.method == 'GET':
        try:
            user_data = CustomUser.objects.filter(id=user_id).values('id', 'username', 'first_name', 'last_name', 'second_last_name', 'email', 'role', 'role__name', 'dining_room_in_charge','dining_room_in_charge__client_diner_dining_room__client__company', 'dining_room_in_charge__name','status', 'created_by', 'updated_by').first()
            return JsonResponse(user_data)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)

            if len(data['username']) < 5:
                return JsonResponse({'error': 'El nombre de usuario debe tener al menos 5 caracteres'}, status=400)
            
            if len(data['first_name']) < 2 or len(data['last_name']) < 2:
                return JsonResponse({'error': 'El nombre y apellido paterno deben tener al menos 2 caracteres'}, status=400)
            
            if 'email' in data and data['email']:
                if '@' not in data['email'] or '.' not in data['email']:
                    return JsonResponse({'error': 'Correo electrónico inválido'}, status=400)
            
            if 'password' in data and data['password'] is not None:
                if len(data['password']) < 8 or not any(char.isdigit() for char in data['password']) or not any(char.isalpha() for char in data['password']):
                    return JsonResponse({'error': 'La contraseña debe tener al menos 8 caracteres, una letra y un número'}, status=400)
                user.password = make_password(data['password'])

            with transaction.atomic():
                user.username = data.get('username', user.username)
                user.email = data.get('email', user.email)
                user.first_name = data.get('first_name', user.first_name)
                user.last_name = data.get('last_name', user.last_name)
                user.second_last_name = data.get('second_last_name', user.second_last_name)
                user.status = data.get('status', user.status)
                if 'role_id' in data:
                    user.role = get_object_or_404(Role, id=data['role_id'])
                user.save()

                if 'dining_room_in_charge' in data and data['dining_room_in_charge'] is not None:
                    if data['dining_room_in_charge'] == 'null' or data['dining_room_in_charge'] == 'no':
                        if user.dining_room_in_charge.exists():
                            dining_room = user.dining_room_in_charge.first()
                            dining_room.in_charge = None
                            dining_room.save()
                        user.dining_room_in_charge.clear()
                    else:
                        dining_room = get_object_or_404(DiningRoom, id=data['dining_room_in_charge'])
                        dining_room.in_charge = user
                        dining_room.save()
                        user.dining_room_in_charge.set([dining_room])

                if data['status'] == False:
                    dining_rooms = DiningRoom.objects.filter(in_charge=user)
                    for dining_room in dining_rooms:
                        dining_room.in_charge = None
                        dining_room.save()
                        user.dining_room_in_charge.clear()
                user.save()

            return JsonResponse({'message': 'Usuario actualizado correctamente'})
        except IntegrityError as e:
            if 'username' in str(e):
                return JsonResponse({'error': 'El nombre de usuario ya existe'}, status=400)
            elif 'email' in str(e):
                return JsonResponse({'error': 'El correo electrónico ya existe'}, status=400)
            else:
                return JsonResponse({'error': 'Error de integridad de datos'}, status=400)
        except KeyError as e:
            return JsonResponse({'error': f'Campo faltante: {str(e)}'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

@csrf_exempt
def role_list(request):
    if request.method == 'GET':
        roles = Role.objects.all().values('id', 'name')
        return JsonResponse(list(roles), safe=False)
    
@csrf_exempt
def get_diner_without_in_charge(request):
    try:
        diners = ClientDiner.objects.filter(dining_room__in_charge__isnull=True, dining_room__status=1).values(
            'dining_room', 'dining_room__name', 'client__company'
        )
        
        # Convert QuerySet to a list of dictionaries
        diners_list = [
            {
                'id': diner['dining_room'],
                'name': diner['dining_room__name'],
                'client_diner_dining_room__client__company': diner['client__company']
            }
            for diner in diners
        ]
        
        return JsonResponse(diners_list, safe=False)
    except Exception as e:
        print(e)
        return JsonResponse({'error': str(e)}, status=500)

# ===================== EMPLEADOS ===================== #
@csrf_exempt
def get_empleados(request):
    page_number = request.GET.get('page', 1)
    page_size = request.GET.get('page_size', 10)
    search_query = request.GET.get('search', '')
    filter_value = request.GET.get('filter', 'all')

    # Filtrar empleados cuyo cliente esté activo
    empleados = Employee.objects.filter(
        employee_client_diner_employee__isnull=False,
        client__status=True  # Solo clientes activos
    ).select_related(
        'client', 'payroll'
    ).prefetch_related(
        'employee_client_diner_employee__client_diner__dining_room'
    ).distinct()

    if search_query:
        empleados = empleados.filter(
            Q(employeed_code__icontains=search_query) |
            Q(name__icontains=search_query) |
            Q(lastname__icontains=search_query) |
            Q(second_lastname__icontains=search_query) |
            Q(client__company__icontains=search_query) |
            Q(payroll__description__icontains=search_query) |
            Q(employee_client_diner_employee__client_diner__dining_room__name__icontains=search_query)
        )

    if filter_value != 'all':
        empleados = empleados.filter(client__id=filter_value)

    empleados = empleados.annotate(
        dining_room_name=models.F('employee_client_diner_employee__client_diner__dining_room__name')
    ).values(
        'id', 'employeed_code', 'name', 'lastname', 'second_lastname', 'client__company', 'dining_room_name', 'payroll__description', 'status'
    ).distinct()

    paginator = Paginator(empleados, page_size)
    page_obj = paginator.get_page(page_number)

    response = {
        'empleados': list(page_obj),
        'total_pages': paginator.num_pages,
        'current_page': page_obj.number,
        'has_previous': page_obj.has_previous(),
        'has_next': page_obj.has_next(),
    }

    return JsonResponse(response)
    

@csrf_exempt
def get_empleado(request):
    try:
        empleado_id = request.GET.get('empleado_id')
        
        # Realizar la consulta con las uniones necesarias
        empleado = Employee.objects.filter(id=empleado_id).select_related(
            'payroll', 'client'
        ).prefetch_related(
            'employee_client_diner_employee__client_diner__dining_room'
        ).values(
            'id',
            'employeed_code',
            'name',
            'lastname',
            'second_lastname',
            'client__company',
            'client_id',
            'payroll__description',
            'payroll_id',
            'status',
            'employee_client_diner_employee__client_diner__dining_room__name',
            'employee_client_diner_employee__client_diner__dining_room__id'
        ).first()

        if not empleado:
            return JsonResponse({'error': 'Empleado no encontrado'}, status=404)

        client = {
            'id': empleado['client_id'],
            'company': empleado['client__company']
        }

        payroll = {
            'id': empleado['payroll_id'],
            'description': empleado['payroll__description']
        }

        dining_room = {
            'id': empleado['employee_client_diner_employee__client_diner__dining_room__id'],
            'name': empleado['employee_client_diner_employee__client_diner__dining_room__name']
        }

        context = {
            'id': empleado['id'],
            'employeed_code': empleado['employeed_code'],
            'name': empleado['name'],
            'lastname': empleado['lastname'],
            'second_lastname': empleado['second_lastname'],
            'client': client,
            'payroll': payroll,
            'status': empleado['status'],
            'dining_room': dining_room
        }

        return JsonResponse(context)
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'Empleado no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def create_empleado(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)            
            
            # Obtener la instancia de PayrollType
            payroll = PayrollType.objects.get(id=data.get('payroll_id'))
            
            # Verificar si el empleado ya existe
            empleado_existente = Employee.objects.filter(employeed_code=data.get('employeed_code')).first()
            
            if empleado_existente:
                # Si el empleado existe y el cliente es el mismo
                if int(empleado_existente.client_id) == int(data.get('client_id')):
                    return JsonResponse({'message': 'El empleado ya existe', 'status': 'danger'}, status=200)
                else:
                    # Si el cliente no es el mismo, crear un nuevo empleado
                    empleado = Employee(
                        employeed_code=data.get('employeed_code'),
                        name=data.get('name').upper(),
                        lastname=data.get('lastname').upper(),
                        second_lastname=data.get('second_lastname').upper(),
                        client_id=data.get('client_id'),
                        payroll=payroll,
                        status=data.get('status'),
                        created_by_id=request.user.id
                    )
                    empleado.save()
            else:
                # Si el empleado no existe, crear un nuevo empleado
                empleado = Employee(
                    employeed_code=data.get('employeed_code'),
                    name=data.get('name').upper(),
                    lastname=data.get('lastname').upper(),
                    second_lastname=data.get('second_lastname').upper(),
                    client_id=data.get('client_id'),
                    payroll=payroll,
                    status=data.get('status'),
                    created_by_id=request.user.id
                )
                empleado.save()

            # Crear la relación en EmployeeClientDiner
            dining_room_id = int(data.get('dining_room_id'))
            if dining_room_id:
                client_diner = ClientDiner.objects.get(client_id=empleado.client_id, dining_room_id=dining_room_id)
                EmployeeClientDiner.objects.create(
                    employee=empleado,
                    client_diner=client_diner,
                    created_by_id=request.user.id,
                    updated_by_id=request.user.id
                )
            
            return JsonResponse({'message': 'Empleado creado correctamente'}, status=201)
        except PayrollType.DoesNotExist:
            return JsonResponse({'error': 'Tipo de nómina no encontrado'}, status=404)
        except ClientDiner.DoesNotExist:
            return JsonResponse({'error': 'Cliente-Comedor no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Método no permitido'}, status=405)


@csrf_exempt
def update_empleado(request):
    try:
        data = json.loads(request.body)
        empleado_id = data.get('id')
        
        # Obtener el empleado existente
        empleado = Employee.objects.get(id=empleado_id)
        
        # Obtener la instancia de PayrollType si se proporciona
        payroll = PayrollType.objects.get(id=data.get('payroll_id')) if data.get('payroll_id') else empleado.payroll
        
        # Verificar si existe un empleado con el mismo código de empleado y cliente
        empleado_existente = Employee.objects.filter(employeed_code=data.get('employeed_code'), client_id=data.get('client_id')).exclude(id=empleado_id).first()
        
        if empleado_existente:
            return JsonResponse({'message': 'El empleado ya existe', 'status': 'danger'}, status=200)
        
        # Actualizar los campos del empleado
        empleado.employeed_code = data.get('employeed_code', empleado.employeed_code)
        empleado.name = data.get('name', empleado.name).upper()
        empleado.lastname = data.get('lastname', empleado.lastname).upper()
        empleado.second_lastname = data.get('second_lastname', empleado.second_lastname).upper()
        empleado.client_id = data.get('client_id', empleado.client_id)
        empleado.payroll = payroll  # Asignar la instancia de PayrollType
        empleado.status = data.get('status', empleado.status)
        
        # Guardar los cambios del empleado
        empleado.save()

        # Actualizar la relación en EmployeeClientDiner
        dining_room_id = data.get('dining_room_id')
        if dining_room_id:
            client_diner = ClientDiner.objects.get(client_id=empleado.client_id, dining_room_id=dining_room_id)
            employee_client_diner, created = EmployeeClientDiner.objects.update_or_create(
                employee=empleado,
                defaults={'client_diner': client_diner, 'updated_by_id': request.user.id}
            )

        return JsonResponse({'message': 'Empleado actualizado correctamente'})
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'Empleado no encontrado'}, status=404)
    except PayrollType.DoesNotExist:
        return JsonResponse({'error': 'Tipo de nómina no encontrado'}, status=404)
    except ClientDiner.DoesNotExist:
        return JsonResponse({'error': 'Cliente-Comedor no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def get_comedores_clientes(request):
    try:
        client_id = request.GET.get('client_id')
        
        if not client_id:
            return JsonResponse({'error': 'El parámetro client_id es obligatorio'}, status=400)
        
        # Realizar la consulta utilizando el ORM de Django
        comedores = DiningRoom.objects.filter(
            client_diner_dining_room__client_id=client_id,
            status=True
        ).values('id', 'name')
        
        comedores_list = list(comedores)
        
        context = {
            'comedores': comedores_list
        }
        
        return JsonResponse(context)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ===================== ARCHIVO EXCEL ===================== #
@csrf_exempt
def upload_empleados(request):
    if request.method == 'POST':
        try:
            # Obtenemos los datos
            data = json.loads(request.body)
            cliente_id = int(data.get('cliente_id'))
            comedor_id = int(data.get('comedor_id'))
            empleados = data.get('empleados')
            
            # Validamos que los campos del modal no estén vacíos
            if not cliente_id:
                return JsonResponse({'error': 'El campo cliente_id es obligatorio'}, status=400)
            if not comedor_id:
                return JsonResponse({'error': 'El campo comedor_id es obligatorio'}, status=400)

            # Contador de empleados insertados, no repetidos y modificados
            empleados_insertados = 0
            empleados_repetidos = 0
            empleados_modificados = 0

            # Iniciamos una transacción al momento de leer el archivo
            with transaction.atomic():
                # Procesamos cada registro en el archivo
                for empleado_data in empleados:
                    # Validamos que los campos obligatorios no estén vacíos
                    if not empleado_data.get('NOMBRES'):
                        return JsonResponse({'error': 'El campo NOMBRES es obligatorio'}, status=400)
                    if not empleado_data.get('NO. EMPLEADO'):
                        return JsonResponse({'error': 'El campo NO. EMPLEADO es obligatorio'}, status=400)
                    if not empleado_data.get('APELLIDO PATERNO'):
                        return JsonResponse({'error': 'El campo APELLIDO PATERNO es obligatorio'}, status=400)
                    if not empleado_data.get('NOMINA'):
                        return JsonResponse({'error': 'El campo NOMINA es obligatorio'}, status=400)

                    # Obtenemos la instancia del modelo PayrollType
                    payroll = PayrollType.objects.get(description=empleado_data.get('NOMINA'))

                    # Asignamos valores predeterminados si los campos son None
                    nombre = empleado_data.get('NOMBRES', '').upper()
                    apellido_paterno = empleado_data.get('APELLIDO PATERNO', '').upper()
                    apellido_materno = empleado_data.get('APELLIDO MATERNO', '').upper()

                    # Obtenemos todos los empleados con el mismo código de empleado
                    empleados_existentes = Employee.objects.filter(employeed_code=empleado_data.get('NO. EMPLEADO'))

                    # Bandera para determinar si el empleado ya existe con los mismos datos
                    empleado_duplicado = False

                    for empleado_existente in empleados_existentes:
                        # Si el cliente es el mismo, verificamos si hay cambios
                        if empleado_existente.client.id == cliente_id:
                            campos_actualizados = False
                            if empleado_existente.name != nombre:
                                empleado_existente.name = nombre
                                campos_actualizados = True
                            if empleado_existente.lastname != apellido_paterno:
                                empleado_existente.lastname = apellido_paterno
                                campos_actualizados = True
                            if empleado_existente.second_lastname != apellido_materno:
                                empleado_existente.second_lastname = apellido_materno
                                campos_actualizados = True
                            if empleado_existente.payroll != payroll:
                                empleado_existente.payroll = payroll
                                campos_actualizados = True
                            if not empleado_existente.status:  # Activamos el registro si está inactivo
                                empleado_existente.status = True
                                campos_actualizados = True

                            # Verificamos y actualizamos la relación con el comedor
                            client_diner = ClientDiner.objects.get(client_id=cliente_id, dining_room_id=comedor_id)
                            employee_client_diner = EmployeeClientDiner.objects.filter(employee=empleado_existente, client_diner__client_id=cliente_id).first()
                            if employee_client_diner:
                                if employee_client_diner.client_diner.dining_room_id != comedor_id:
                                    employee_client_diner.client_diner = client_diner
                                    employee_client_diner.updated_by_id = request.user.id
                                    employee_client_diner.save()
                                    campos_actualizados = True

                            if campos_actualizados:
                                empleado_existente.updated_by_id = request.user.id
                                empleado_existente.save()
                                empleados_modificados += 1  # Incrementar solo una vez si hubo cambios
                            else:
                                empleados_repetidos += 1  # Si no hubo cambios, se considera repetido
                            empleado_duplicado = True
                            break

                    # Si no se encontró un duplicado exacto, se inserta un nuevo empleado
                    if not empleado_duplicado:
                        empleado = Employee(
                            employeed_code=empleado_data.get('NO. EMPLEADO'),
                            name=nombre,
                            lastname=apellido_paterno,
                            second_lastname=apellido_materno,
                            client_id=cliente_id,
                            payroll=payroll,
                            status=empleado_data.get('ESTADO', True),
                            created_by_id=request.user.id
                        )
                        empleado.save()
                        empleados_insertados += 1

                        # Creamos la relación en el modelo EmployeeClientDiner
                        client_diner = ClientDiner.objects.get(client_id=cliente_id, dining_room_id=comedor_id)
                        EmployeeClientDiner.objects.create(
                            employee=empleado,
                            client_diner=client_diner,
                            created_by_id=request.user.id,
                            updated_by_id=request.user.id
                        )

            # Mensajes de respuesta
            message = {}

            if empleados_insertados > 0:
                message['message1'] = [f'Empleados insertados: {empleados_insertados}', 'success']
            if empleados_repetidos > 0:
                message['message2'] = [f'Empleados ya existentes: {empleados_repetidos}', 'info']
            if empleados_modificados > 0:
                message['message3'] = [f'Empleados modificados: {empleados_modificados}', 'info']

            return JsonResponse({'message': message})
        except PayrollType.DoesNotExist:
            return JsonResponse({'error': 'Tipo de nómina no encontrado'}, status=404)
        except ClientDiner.DoesNotExist:
            return JsonResponse({'error': 'Cliente-Comedor no encontrado'}, status=404)
        except DiningRoom.DoesNotExist:
            return JsonResponse({'error': 'Comedor no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Método no permitido'}, status=405)
        

# ===================== TIPOS DE NOMINA ===================== #
@csrf_exempt
def get_tipos_nomina(request):
    try:
        tipos_nomina = PayrollType.objects.all().values('id', 'description')
        context = {
            'tipos_nomina': list(tipos_nomina)
        }
        return JsonResponse(context)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ===================== CLIENTES ===================== #
@csrf_exempt
def get_clientes(request):
    try:
        clientes = Client.objects.filter(status=1).values('id', 'company')
        context = {
            'clientes': list(clientes)
        }        
        return JsonResponse(context)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# ===================== REPORTE EMPLEADOS ===================== #
@csrf_exempt
def get_employee_report_general(request):
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    page = request.GET.get('page', 1)
    page_size = request.GET.get('page_size', 10)

    try:
        filters = {
            'employee_client_diner__client_diner__client__id': request.GET.get('filterClient'),
            'employee_client_diner__client_diner__dining_room__id': request.GET.get('filterDiningRoom'),
            'employee_client_diner__employee__employeed_code__icontains': request.GET.get('filterEmployeeNumber'),
            'employee_client_diner__employee__status': request.GET.get('filterStatus'),
            'created_at__gte': request.GET.get('filterStartDate'),
            'created_at__lte': request.GET.get('filterEndDate'),
            'employee_client_diner__isnull': False,
            'voucher__isnull': True
        }

        filters = {k: v for k, v in filters.items() if v}

        entry_employee = Entry.objects.select_related(
            'employee_client_diner__employee',
            'employee_client_diner__client_diner__client',
            'employee_client_diner__client_diner__dining_room'
        ).filter(**filters).values(
            client_company=F('employee_client_diner__client_diner__client__company'),
            client_name=F('employee_client_diner__client_diner__client__name'),
            client_lastname=F('employee_client_diner__client_diner__client__lastname'),
            client_second_lastname=F('employee_client_diner__client_diner__client__second_lastname'),
            dining_room_name=F('employee_client_diner__client_diner__dining_room__name'),
            employee_code=F('employee_client_diner__employee__employeed_code'),
            employee_name=F('employee_client_diner__employee__name'),
            employee_lastname=F('employee_client_diner__employee__lastname'),
            employee_second_lastname=F('employee_client_diner__employee__second_lastname'),
            employee_status=F('employee_client_diner__employee__status'),
            entry_created_at=F('created_at')
        ).order_by('-created_at')

        paginator = Paginator(entry_employee, page_size)

        try:
            entry_employee = paginator.page(page)
        except PageNotAnInteger:
            entry_employee = paginator.page(1)
        except EmptyPage:
            entry_employee = paginator.page(paginator.num_pages)

        context = {
            'entry_employee': list(entry_employee),
            'page': entry_employee.number,
            'pages': paginator.num_pages,
            'has_previous': entry_employee.has_previous(),
            'has_next': entry_employee.has_next()
        }

        return JsonResponse(context)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def get_clients_employee_reports(request):
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        filters = {
            'employee_client_diner__client_diner__client__id': request.GET.get('filterClient'),
            'employee_client_diner__client_diner__dining_room__id': request.GET.get('filterDiningRoom'),
            'employee_client_diner__employee__employeed_code__icontains': request.GET.get('filterEmployeeNumber'),
            'employee_client_diner__employee__status': request.GET.get('filterStatus'),
            'created_at__gte': request.GET.get('filterStartDate'),
            'created_at__lte': request.GET.get('filterEndDate'),
            'employee_client_diner__isnull': False,
            'voucher__isnull': True
        }
        filters = {k: v for k, v in filters.items() if v}

        clients = Entry.objects.select_related(
            'employee_client_diner__client_diner__client'
        ).filter(**filters).values(
            'employee_client_diner__client_diner__client__id',
            'employee_client_diner__client_diner__client__company'
        ).distinct()

        context = {
            'clients': list(clients)
        }

        return JsonResponse(context)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def get_diner_employee_reports(request):
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        filters = {
            'employee_client_diner__client_diner__client__id': request.GET.get('filterClient'),
            'employee_client_diner__client_diner__dining_room__id': request.GET.get('filterDiningRoom'),
            'employee_client_diner__employee__employeed_code__icontains': request.GET.get('filterEmployeeNumber'),
            'employee_client_diner__employee__status': request.GET.get('filterStatus'),
            'created_at__gte': request.GET.get('filterStartDate'),
            'created_at__lte': request.GET.get('filterEndDate'),
            'employee_client_diner__isnull': False,
            'voucher__isnull': True
        }
        filters = {k: v for k, v in filters.items() if v}

        diners = Entry.objects.select_related(
            'employee_client_diner__client_diner__dining_room'
        ).filter(**filters).values(
            'employee_client_diner__client_diner__dining_room__id',
            'employee_client_diner__client_diner__dining_room__name'
        ).distinct()

        context = {
            'diners': list(diners)
        }

        return JsonResponse(context)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def get_employee_report_summary(request):
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    page_number = request.GET.get('page', 1)
    page_size = request.GET.get('page_size', 10)
    try:
        filters = {
            'employee_client_diner__client_diner__client__id': request.GET.get('filterClient'),
            'employee_client_diner__client_diner__dining_room__id': request.GET.get('filterDiningRoom'),
            'employee_client_diner__employee__employeed_code__icontains': request.GET.get('filterEmployeeNumber'),
            'employee_client_diner__employee__status': request.GET.get('filterStatus'),
            'created_at__gte': request.GET.get('filterStartDate'),
            'created_at__lte': request.GET.get('filterEndDate'),
            'employee_client_diner__isnull': False,
            'voucher__isnull': True
        }
        filters = {k: v for k, v in filters.items() if v}

        employee_report_summary = Entry.objects.select_related(
            'employee_client_diner__employee',
            'employee_client_diner__client_diner__client',
            'employee_client_diner__client_diner__dining_room'
        ).filter(**filters).values(
            employee_id=F('employee_client_diner__employee__id'),
            dining_room_id=F('employee_client_diner__client_diner__dining_room__id'),
            client_company=F('employee_client_diner__client_diner__client__company'),
            client_name=F('employee_client_diner__client_diner__client__name'),
            client_lastname=F('employee_client_diner__client_diner__client__lastname'),
            client_second_lastname=F('employee_client_diner__client_diner__client__second_lastname'),
            dining_room_name=F('employee_client_diner__client_diner__dining_room__name'),
            employee_code=F('employee_client_diner__employee__employeed_code'),
            employee_name=F('employee_client_diner__employee__name'),
            employee_lastname=F('employee_client_diner__employee__lastname'),
            employee_second_lastname=F('employee_client_diner__employee__second_lastname'),
            employee_status=F('employee_client_diner__employee__status')
        ).annotate(
            entry_count=Count('id')
        ).order_by('employee_code')

        paginator = Paginator(employee_report_summary, page_size)
        try:
            employee_report_summary = paginator.page(page_number)
        except PageNotAnInteger:
            employee_report_summary = paginator.page(1)
        except EmptyPage:
            employee_report_summary = paginator.page(paginator.num_pages)

        context = {
            'employee_report_summary': list(employee_report_summary),
            'page': employee_report_summary.number,
            'pages': paginator.num_pages,
            'has_previous': employee_report_summary.has_previous(),
            'has_next': employee_report_summary.has_next()
        }

        return JsonResponse(context)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def get_employee_report_summary_details(request):
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    employee_id = request.GET.get('employeeId')
    diner_id = request.GET.get('dinerId')
    page_number = request.GET.get('page', 1)
    page_size = request.GET.get('page_size', 5)

    if not employee_id:
        return JsonResponse({'error': 'employee_id es requerido'}, status=400)
    
    try:
        filters = {
            'employee_client_diner__client_diner__client__id': request.GET.get('filterClient'),
            'employee_client_diner__client_diner__dining_room__id': request.GET.get('filterDiningRoom'),
            'employee_client_diner__employee__employeed_code__icontains': request.GET.get('filterEmployeeNumber'),
            'employee_client_diner__employee__status': request.GET.get('filterStatus'),
            'created_at__gte': request.GET.get('filterStartDate'),
            'created_at__lte': request.GET.get('filterEndDate'),
            'employee_client_diner__employee__id': employee_id,
            'lots__voucher_type_id': None
        }
        filters = {k: v for k, v in filters.items() if v}

        employee_detail = Entry.objects.select_related(
            'employee_client_diner__employee',
            'employee_client_diner__client_diner__client',
            'employee_client_diner__client_diner__dining_room'
        ).filter(**filters).values(
            employee_code=F('employee_client_diner__employee__employeed_code'),
            employee_name=F('employee_client_diner__employee__name'),
            employee_lastname=F('employee_client_diner__employee__lastname'),
            employee_second_lastname=F('employee_client_diner__employee__second_lastname'),
            employee_status=F('employee_client_diner__employee__status'),
        ).annotate(
            entry=Count('id')
        ).first()

        filters['employee_client_diner__client_diner__dining_room__id'] = diner_id
        
        # Traer las entradas del empleado
        employee_entries = Entry.objects.select_related(
            'employee_client_diner__client_diner__client',
            'employee_client_diner__client_diner__dining_room'
        ).filter(**filters).values(
            client_company=F('employee_client_diner__client_diner__client__company'),
            client_name=F('employee_client_diner__client_diner__client__name'),
            client_lastname=F('employee_client_diner__client_diner__client__lastname'),
            client_second_lastname=F('employee_client_diner__client_diner__client__second_lastname'),
            dining_room_name=F('employee_client_diner__client_diner__dining_room__name'),
            entry_created_at=F('created_at')
        ).order_by('-created_at')

        employee_entries_len = len(employee_entries)
        employee_detail['entry_count'] = employee_entries_len

        paginator = Paginator(employee_entries, page_size)
        try:
            employee_entries = paginator.page(page_number)
        except PageNotAnInteger:
            employee_entries = paginator.page(1)
        except EmptyPage:
            employee_entries = paginator.page(paginator.num_pages)

        context = {
            'employee_detail': employee_detail,
            'employee_entries': list(employee_entries),
            'page': employee_entries.number,
            'pages': paginator.num_pages,
            'has_previous': employee_entries.has_previous(),
            'has_next': employee_entries.has_next()
        }

        return JsonResponse(context)
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'Empleado no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def export_excel_employee_report(request):
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    try:
        filters = {
            'employee_client_diner__client_diner__client__id': request.GET.get('filterClient'),
            'employee_client_diner__client_diner__dining_room__id': request.GET.get('filterDiningRoom'),
            'employee_client_diner__employee__employeed_code__icontains': request.GET.get('filterEmployeeNumber'),
            'employee_client_diner__employee__status': request.GET.get('filterStatus'),
            'employee_client_diner__isnull': False,
            'voucher__isnull': True
        }

        # Convertir las fechas a objetos datetime conscientes de la zona horaria
        start_date_str = request.GET.get('filterStartDate')
        end_date_str = request.GET.get('filterEndDate')

        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            start_date = timezone.make_aware(start_date)
        else:
            start_date = None

        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            end_date = timezone.make_aware(end_date)
        else:
            end_date = None

        # Filtro de entradas solo cuando se han proporcionado fechas
        if start_date:
            filters['created_at__gte'] = start_date
        if end_date:
            filters['created_at__lte'] = end_date

        filters = {k: v for k, v in filters.items() if v}

        entry_employee = Entry.objects.select_related(
            'employee_client_diner__employee',
            'employee_client_diner__employee__payroll',
            'employee_client_diner__client_diner__client',
            'employee_client_diner__client_diner__dining_room'
        ).filter(**filters).values(
            client_company=F('employee_client_diner__client_diner__client__company'),
            client_name=F('employee_client_diner__client_diner__client__name'),
            client_lastname=F('employee_client_diner__client_diner__client__lastname'),
            client_second_lastname=F('employee_client_diner__client_diner__client__second_lastname'),
            dining_room_name=F('employee_client_diner__client_diner__dining_room__name'),
            employee_code=F('employee_client_diner__employee__employeed_code'),
            employee_name=F('employee_client_diner__employee__name'),
            employee_lastname=F('employee_client_diner__employee__lastname'),
            employee_second_lastname=F('employee_client_diner__employee__second_lastname'),
            employee_status=F('employee_client_diner__employee__status'),
            entry_created_at=F('created_at')
        ).order_by('-created_at')            

        # Crear un DataFrame a partir del queryset
        df = pd.DataFrame(list(entry_employee))

        # Eliminar la información de la zona horaria de los objetos datetime
        if 'entry_created_at' in df.columns:
            df['entry_created_at'] = df['entry_created_at'].apply(lambda x: format_date(x) if pd.notnull(x) else "Sin fecha")
        
        # Si el estado del empleado es True, cambiar a 'Activo', de lo contrario 'Inactivo'
        df['employee_status'] = df['employee_status'].apply(lambda x: 'Activo' if x else 'Inactivo')
        

        # Definir los encabezados personalizados para la hoja de Excel
        headers = [
            'Compañía Cliente',
            'Nombre Cliente',
            'Apellido Paterno Cliente',
            'Apellido Materno Cliente',
            'Nombre Comedor',
            'Código Empleado',
            'Nombre Empleado',
            'Apellido Paterno Empleado',
            'Apellido Materno Empleado',
            'Estado Empleado',
            'Fecha de Ingreso'
        ]

        # Crear un nuevo libro de Excel y agregar dos hojas
        wb = openpyxl.Workbook()

        # Hoja 1: Informe de empleados
        ws1 = wb.active
        ws1.title = "Employee Report"
        ws1.append(headers)
        add_styles(ws1, headers)

        for row in dataframe_to_rows(df, index=False, header=False):
            ws1.append(row)
        
        # Hoja 2: Informe resumido de empleados
        ws2 = wb.create_sheet(title="Summary Report")
        simplified_headers = [
            'Código Empleado',
            'Nombre Empleado',
            'Apellido Paterno Empleado',
            'Apellido Materno Empleado',
            'Estado Empleado',
            'Nombre Comedor',
            'Total Entradas',
            'Nomina'
        ]
        ws2.append(simplified_headers)
        add_styles(ws2, simplified_headers)

        # Generar un informe resumido (agregando por empleado)
        summary_df = df.groupby(['employee_code', 'employee_name', 'employee_lastname', 'employee_second_lastname', 'dining_room_name', 'employee_status']).size().reset_index(name='Total Entradas')


        def add_payroll_types(row):
            employee = Employee.objects.filter(employeed_code=row['employee_code']).first()
            print(employee.payroll)
            row['payroll_description'] = employee.payroll.description
            return row
    
        summary_df = summary_df.apply(add_payroll_types, axis=1)

        for row in dataframe_to_rows(summary_df, index=False, header=False):
            ws2.append(row)

        # Crear un buffer de BytesIO para contener el archivo de Excel
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Establecer el tipo de contenido de la respuesta a Excel
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=employee_report.xlsx'

        return response
    except Exception as e:
        return JsonResponse({'error': 'No hay registros'}, status=500)

def add_styles(ws, headers):
    # Set font and style for headers
    bold_font = Font(bold=True)
    fill_color = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = bold_font
        cell.fill = fill_color
        cell.border = border
        cell.value = header

    # Apply border to all cells in the sheet
    for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = border


# ===================== REPORTE VALES UNICOS ===================== #
@csrf_exempt
def get_unique_reports(request):
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    page_number = request.GET.get('page', 1)
    page_size = request.GET.get('page_size', 10)
    lot = request.GET.get('filterLotNumber')


    try:
        # Filtros de vales
        filters = {
            'lots__client_diner__client__id': request.GET.get('filterClient'),
            'lots__client_diner__dining_room__id': request.GET.get('filterDiningRoom'),
            'folio__icontains': request.GET.get('filterVoucherNumber'),
            'status': request.GET.get('filterStatus'),
            'lots__voucher_type_id': 1,
            'lots__id': lot
        }

        filters = {k: v for k, v in filters.items() if v}

        # Convertir las fechas a datetime si están presentes
        start_date = request.GET.get('filterStartDate')
        end_date = request.GET.get('filterEndDate')
        
        if start_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        if end_date:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

        # Filtro de entradas solo cuando se han proporcionado fechas
        entry_filter = {}
        if start_date:
            entry_filter['created_at__gte'] = start_date
        if end_date:
            entry_filter['created_at__lte'] = end_date

        # Si se especifica alguna fecha, se filtran solo los vales con entradas dentro del rango
        if entry_filter:
            vouchers = Voucher.objects.select_related(
                'lots__client_diner__client',
                'lots__client_diner__dining_room',
            ).filter(**filters).filter(
                Exists(
                    Entry.objects.filter(voucher_id=OuterRef('id')).filter(**entry_filter)
                )
            ).values(
                'id',
                client_company=F('lots__client_diner__client__company'),
                client_name=F('lots__client_diner__client__name'),
                client_lastname=F('lots__client_diner__client__lastname'),
                client_second_lastname=F('lots__client_diner__client__second_lastname'),
                dining_room_name=F('lots__client_diner__dining_room__name'),
                voucher_folio=F('folio'),
                voucher_status=F('status')
            ).order_by('-id')
        else:
            # Si no hay fechas, no aplicamos el filtro de entradas
            vouchers = Voucher.objects.select_related(
                'lots__client_diner__client',
                'lots__client_diner__dining_room',
            ).filter(**filters).values(
                'id',
                client_company=F('lots__client_diner__client__company'),
                client_name=F('lots__client_diner__client__name'),
                client_lastname=F('lots__client_diner__client__lastname'),
                client_second_lastname=F('lots__client_diner__client__second_lastname'),
                dining_room_name=F('lots__client_diner__dining_room__name'),
                voucher_folio=F('folio'),
                voucher_status=F('status')
            ).order_by('-id')

        # Obtener los IDs de los vales para buscar las entradas
        voucher_ids = [voucher['id'] for voucher in vouchers]
        
        # Obtener las entradas dentro del rango de fechas si existen fechas
        entry_filters = {
            'voucher_id__in': voucher_ids
        }
        if start_date:
            entry_filters['created_at__gte'] = start_date
        if end_date:
            entry_filters['created_at__lte'] = end_date

        entries = Entry.objects.filter(**entry_filters).values(
            'voucher_id',
            'created_at'
        )

        # Crear un diccionario para mapear las entradas a los vales
        entry_map = {entry['voucher_id']: entry['created_at'] for entry in entries}

        # Añadir la información de las entradas a los vales
        for voucher in vouchers:
            voucher['entry_created_at'] = entry_map.get(voucher['id'], None)

        paginator = Paginator(vouchers, page_size)

        try:
            vouchers = paginator.page(page_number)
        except PageNotAnInteger:
            vouchers = paginator.page(1)
        except EmptyPage:
            vouchers = paginator.page(paginator.num_pages)

        context = {
            'unique_reports': list(vouchers),
            'page': vouchers.number,
            'pages': paginator.num_pages,
            'has_previous': vouchers.has_previous(),
            'has_next': vouchers.has_next()
        }

        return JsonResponse(context)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
@csrf_exempt
def get_clients_unique_reports(request):
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    lot = request.GET.get('filterLotNumber')

    try:
        # Filtros para los vales
        filters = {
            'lots__client_diner__client__id': request.GET.get('filterClient'),
            'lots__client_diner__dining_room__id': request.GET.get('filterDiningRoom'),
            'folio__icontains': request.GET.get('filterVoucherNumber'),
            'status': request.GET.get('filterStatus'),
            'lots__voucher_type_id': 1,
            'lots__id': lot
        }
        filters = {k: v for k, v in filters.items() if v}

        # Convertir las fechas a datetime si están presentes
        start_date = request.GET.get('filterStartDate')
        end_date = request.GET.get('filterEndDate')
        
        if start_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        if end_date:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

        # Filtrar los vales
        vouchers = Voucher.objects.select_related(
            'lots__client_diner__client'
        ).filter(**filters)

        # Filtrar los clientes únicos de los vales
        clients = vouchers.values(
            client_id=F('lots__client_diner__client__id'),
            client_company=F('lots__client_diner__client__company')
        ).distinct()

        context = {
            'clients': list(clients)
        }

        return JsonResponse(context)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def get_diners_unique_reports(request):
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    lot = request.GET.get('filterLotNumber')

    
    try:
        # Filtros para los vales
        filters = {
            'lots__client_diner__client__id': request.GET.get('filterClient'),
            'lots__client_diner__dining_room__id': request.GET.get('filterDiningRoom'),
            'folio__icontains': request.GET.get('filterVoucherNumber'),
            'status': request.GET.get('filterStatus'),
            'lots__voucher_type_id': 1,
            'lots__id': lot
        }
        filters = {k: v for k, v in filters.items() if v}

        # Convertir las fechas a datetime si están presentes
        start_date = request.GET.get('filterStartDate')
        end_date = request.GET.get('filterEndDate')
        
        if start_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        if end_date:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

        # Filtrar los vales
        vouchers = Voucher.objects.select_related(
            'lots__client_diner__dining_room'
        ).filter(**filters)

        # Filtrar los comensales únicos de los vales
        diners = vouchers.values(
            diner_id=F('lots__client_diner__dining_room__id'),
            diner_name=F('lots__client_diner__dining_room__name')
        ).distinct()

        context = {
            'diners': list(diners)
        }

        return JsonResponse(context)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
@csrf_exempt
def export_excel_unique_reports(request):
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    lot = request.GET.get('filterLotNumber')

    try:
        # Filtros de vales
        filters = {
            'lots__client_diner__client__id': request.GET.get('filterClient'),
            'lots__client_diner__dining_room__id': request.GET.get('filterDiningRoom'),
            'folio__icontains': request.GET.get('filterVoucherNumber'),
            'status': request.GET.get('filterStatus'),
            'lots__voucher_type_id': 1,
            'lots__id': lot
        }

        filters = {k: v for k, v in filters.items() if v}

        # Convertir las fechas a datetime si están presentes
        start_date = request.GET.get('filterStartDate')
        end_date = request.GET.get('filterEndDate')
        
        if start_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        if end_date:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

        # Filtro de entradas solo cuando se han proporcionado fechas
        entry_filter = {}
        if start_date:
            entry_filter['created_at__gte'] = start_date
        if end_date:
            entry_filter['created_at__lte'] = end_date

        # Si se especifica alguna fecha, se filtran solo los vales con entradas dentro del rango
        if entry_filter:
            vouchers = Voucher.objects.select_related(
                'lots__client_diner__client',
                'lots__client_diner__dining_room',
            ).filter(**filters).filter(
                Exists(
                    Entry.objects.filter(voucher_id=OuterRef('id')).filter(**entry_filter)
                )
            ).values(
                'id',
                client_company=F('lots__client_diner__client__company'),
                client_name=F('lots__client_diner__client__name'),
                client_lastname=F('lots__client_diner__client__lastname'),
                client_second_lastname=F('lots__client_diner__client__second_lastname'),
                dining_room_name=F('lots__client_diner__dining_room__name'),
                voucher_folio=F('folio'),
                voucher_status=F('status')
            ).order_by('-id')
        else:
            # Si no hay fechas, no aplicamos el filtro de entradas
            vouchers = Voucher.objects.select_related(
                'lots__client_diner__client',
                'lots__client_diner__dining_room',
            ).filter(**filters).values(
                'id',
                client_company=F('lots__client_diner__client__company'),
                client_name=F('lots__client_diner__client__name'),
                client_lastname=F('lots__client_diner__client__lastname'),
                client_second_lastname=F('lots__client_diner__client__second_lastname'),
                dining_room_name=F('lots__client_diner__dining_room__name'),
                voucher_folio=F('folio'),
                voucher_status=F('status')
            ).order_by('-id')

        # Obtener los IDs de los vales para buscar las entradas
        voucher_ids = [voucher['id'] for voucher in vouchers]
        
        # Obtener las entradas dentro del rango de fechas si existen fechas
        entry_filters = {
            'voucher_id__in': voucher_ids
        }
        if start_date:
            entry_filters['created_at__gte'] = start_date
        if end_date:
            entry_filters['created_at__lte'] = end_date

        entries = Entry.objects.filter(**entry_filters).values(
            'voucher_id',
            'created_at'
        )

        # Crear un diccionario para mapear las entradas a los vales
        entry_map = {entry['voucher_id']: entry['created_at'] for entry in entries}

        # Añadir la información de las entradas a los vales
        for voucher in vouchers:
            voucher['entry_created_at'] = entry_map.get(voucher['id'], None)

        # Create a DataFrame from the queryset
        df = pd.DataFrame(list(vouchers))

        # Remove the 'id' column
        df.drop(columns=['id'], inplace=True)

        # Remove timezone information from datetime objects and handle null values
        if 'entry_created_at' in df.columns:
            df['entry_created_at'] = df['entry_created_at'].apply(lambda x: format_date(x) if pd.notnull(x) else "Sin usar")

        # Si el estado del vale es True, cambiar a 'Activo', de lo contrario 'Inactivo'
        df['voucher_status'] = df['voucher_status'].apply(lambda x: 'Sin usar' if x else 'Usado')

        # Define the custom headers for the Excel sheet
        headers = [
            'Compañía Cliente',
            'Nombre Cliente',
            'Apellido Paterno Cliente',
            'Apellido Materno Cliente',
            'Nombre Comedor',
            'Folio Vale',
            'Fecha de Uso',
            'Estado Vale'
        ]

        # Reorder the DataFrame columns to match the headers
        df = df[['client_company', 'client_name', 'client_lastname', 'client_second_lastname', 'dining_room_name', 'voucher_folio', 'entry_created_at', 'voucher_status']]

        # Create a new Excel workbook and add two sheets
        wb = openpyxl.Workbook()

        # Sheet 1: Unique Voucher Report
        ws1 = wb.active
        ws1.title = "Unique Voucher Report"
        ws1.append(headers)
        add_styles(ws1, headers)

        for row in dataframe_to_rows(df, index=False, header=False):
            ws1.append(row)

        # Create a BytesIO buffer to hold the Excel file
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Set the response content type to Excel
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=unique_voucher_report.xlsx'

        return response
    except Exception as e:
        return JsonResponse({'error': 'No hay registros'}, status=500)

def format_date(date):
    if not date:
        return "Sin usar"
    date_obj = date.replace(tzinfo=None) - timedelta(hours=7)
    day = date_obj.day
    month = date_obj.month
    year = date_obj.year
    hours = date_obj.hour
    minutes = date_obj.minute
    seconds = date_obj.second
    return f"{day:02d}/{month:02d}/{year} - {hours:02d}:{minutes:02d}:{seconds:02d}"

# ===================== REPORTE VALES PERPETUOS ===================== #
@csrf_exempt
def get_perpetual_reports(request):
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    page_number = request.GET.get('page', 1)
    page_size = request.GET.get('page_size', 10)
    lot = request.GET.get('filterLotNumber')


    try:
        # Filtros de vales
        filters = {
            'voucher__lots__client_diner__client__id': request.GET.get('filterClient'),
            'voucher__lots__client_diner__dining_room__id': request.GET.get('filterDiningRoom'),
            'employee__icontains': request.GET.get('filterEmployeeName'),
            'voucher__status': request.GET.get('filterStatus'),
            'voucher__lots__voucher_type_id': 2,  # Tipo de vale perpetuo
            'voucher__folio__icontains': request.GET.get('filterVoucherFolio'),
            'voucher__lots__id': lot
        }

        filters = {k: v for k, v in filters.items() if v}

        # Convertir las fechas a datetime si están presentes
        start_date = request.GET.get('filterStartDate')
        end_date = request.GET.get('filterEndDate')

        if start_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        if end_date:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

        # Filtro de entradas solo cuando se han proporcionado fechas
        entry_filter = {}
        if start_date:
            entry_filter['created_at__gte'] = start_date
        if end_date:
            entry_filter['created_at__lte'] = end_date

        # Filtrar solo las entradas cuyo vale es de tipo 1
        entries = Entry.objects.select_related(
            'voucher__lots__client_diner__client',  # Corrected
            'voucher__lots__client_diner__dining_room'  # Corrected
        ).filter(
            **entry_filter
        ).filter(**filters).values(
            'voucher_id',
            client_company=F('voucher__lots__client_diner__client__company'),
            client_name=F('voucher__lots__client_diner__client__name'),
            client_lastname=F('voucher__lots__client_diner__client__lastname'),
            client_second_lastname=F('voucher__lots__client_diner__client__second_lastname'),
            dining_room_name=F('voucher__lots__client_diner__dining_room__name'),
            voucher_folio=F('voucher__folio'),
            employee_name=F('voucher__employee'),
            voucher_status=F('voucher__status'),
            entry_created_at=F('created_at')
        ).order_by('-created_at')

        # Pagination and response handling
        paginator = Paginator(entries, page_size)

        try:
            entries = paginator.page(page_number)
        except PageNotAnInteger:
            entries = paginator.page(1)
        except EmptyPage:
            entries = paginator.page(paginator.num_pages)

        context = {
            'perpetual_reports': list(entries),
            'page': entries.number,
            'pages': paginator.num_pages,
            'has_previous': entries.has_previous(),
            'has_next': entries.has_next()
        }
        return JsonResponse(context)
    except Exception as e:
        print(e)
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def get_clients_perpetual_reports(request):
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    lot = request.GET.get('filterLotNumber')


    try:
        # Filtros para los vales
        filters = {
            'lots__client_diner__client__id': request.GET.get('filterClient'),
            'lots__client_diner__dining_room__id': request.GET.get('filterDiningRoom'),
            'employee__icontains': request.GET.get('filterEmployeeName'),
            'status': request.GET.get('filterStatus'),
            'lots__voucher_type_id': 2,  # Tipo de vale perpetuo
            'folio__icontains': request.GET.get('filterVoucherFolio'),
            'lots__id': lot
        }
        filters = {k: v for k, v in filters.items() if v}

        # Convertir las fechas a datetime si están presentes
        start_date = request.GET.get('filterStartDate')
        end_date = request.GET.get('filterEndDate')

        if start_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        if end_date:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

        # Filtrar los vales de tipo 1
        vouchers = Voucher.objects.select_related(
            'lots__client_diner__client'
        ).filter(**filters)

        # Filtrar los clientes únicos de los vales de tipo 1
        clients = vouchers.values(
            client_id=F('lots__client_diner__client__id'),
            client_company=F('lots__client_diner__client__company')
        ).distinct()

        context = {
            'clients': list(clients)
        }

        return JsonResponse(context)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def get_diners_perpetual_reports(request):
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    lot = request.GET.get('filterLotNumber')


    try:
        # Filtros para los vales
        filters = {
            'lots__client_diner__client__id': request.GET.get('filterClient'),
            'lots__client_diner__dining_room__id': request.GET.get('filterDiningRoom'),
            'employee__icontains': request.GET.get('filterEmployeeName'),
            'status': request.GET.get('filterStatus'),
            'lots__voucher_type_id': 2,  # Tipo de vale perpetuo
            'folio__icontains': request.GET.get('filterVoucherFolio'),
            'lots__id': lot
        }
        filters = {k: v for k, v in filters.items() if v}

        # Convertir las fechas a datetime si están presentes
        start_date = request.GET.get('filterStartDate')
        end_date = request.GET.get('filterEndDate')

        if start_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        if end_date:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

        # Filtrar los vales de tipo 1
        vouchers = Voucher.objects.select_related(
            'lots__client_diner__dining_room'
        ).filter(**filters)

        # Filtrar los comensales únicos de los vales de tipo 1
        diners = vouchers.values(
            diner_id=F('lots__client_diner__dining_room__id'),
            diner_name=F('lots__client_diner__dining_room__name')
        ).distinct()

        context = {
            'diners': list(diners)
        }

        return JsonResponse(context)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def get_perpetual_report_summary(request):
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    page_number = request.GET.get('page', 1)
    page_size = request.GET.get('page_size', 10)
    lot = request.GET.get('filterLotNumber')

    
    try:
        filters = {
            'lots__client_diner__client__id': request.GET.get('filterClient'),
            'lots__client_diner__dining_room__id': request.GET.get('filterDiningRoom'),
            'employee__icontains': request.GET.get('filterEmployeeName'),
            'status': request.GET.get('filterStatus'),
            'lots__voucher_type_id': 2,  # Tipo de vale perpetuo
            'folio__icontains': request.GET.get('filterVoucherFolio'),
            'lots__id': lot
        }
        filters = {k: v for k, v in filters.items() if v}

        # Obtener los vales de tipo 1
        perpetual_report_summary = Voucher.objects.select_related(
            'lots__client_diner__client',
            'lots__client_diner__dining_room'
        ).filter(**filters).values(
            voucher_id=F('id'),
            client_company=F('lots__client_diner__client__company'),
            client_name=F('lots__client_diner__client__name'),
            client_lastname=F('lots__client_diner__client__lastname'),
            client_second_lastname=F('lots__client_diner__client__second_lastname'),
            dining_room_name=F('lots__client_diner__dining_room__name'),
            voucher_folio=F('folio'),
            employee_name=F('employee'),
            voucher_status=F('status')
        ).annotate(
            entry_count=Count('entry_voucher')
        ).order_by('voucher_folio')

        paginator = Paginator(perpetual_report_summary, page_size)
        try:
            perpetual_report_summary = paginator.page(page_number)
        except PageNotAnInteger:
            perpetual_report_summary = paginator.page(1)
        except EmptyPage:
            perpetual_report_summary = paginator.page(paginator.num_pages)

        context = {
            'perpetual_report_summary': list(perpetual_report_summary),
            'page': perpetual_report_summary.number,
            'pages': paginator.num_pages,
            'has_previous': perpetual_report_summary.has_previous(),
            'has_next': perpetual_report_summary.has_next()
        }

        return JsonResponse(context)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def get_perpetual_report_summary_details(request):
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    voucher_id = request.GET.get('voucherId')
    page_number = request.GET.get('page', 1)
    page_size = request.GET.get('page_size', 5)
    lot = request.GET.get('filterLotNumber')

    if not voucher_id:
        return JsonResponse({'error': 'voucher_id es requerido'}, status=400)
    
    try:
        filters = {
            'id': voucher_id,
            'lots__id': lot,
        }

        voucher_detail = Voucher.objects.select_related(
            'lots__client_diner__client',
            'lots__client_diner__dining_room'
        ).filter(**filters).values(
            voucher_folio=F('folio'),
            employee_name=F('employee'),
            voucher_status=F('status'),
        ).annotate(
            entry=Count('entry_voucher')
        ).first()

        # Traer las entradas del vale
        entry_filters = {
            'voucher_id': voucher_id
        }

        voucher_entries = Entry.objects.select_related(
            'client_diner__client',
            'client_diner__dining_room'
        ).filter(**entry_filters).values(
            client_company=F('client_diner__client__company'),
            client_name=F('client_diner__client__name'),
            client_lastname=F('client_diner__client__lastname'),
            client_second_lastname=F('client_diner__client__second_lastname'),
            dining_room_name=F('client_diner__dining_room__name'),
            entry_created_at=F('created_at')
        ).order_by('-created_at')

        voucher_entries_len = len(voucher_entries)
        voucher_detail['entry_count'] = voucher_entries_len

        paginator = Paginator(voucher_entries, page_size)
        try:
            voucher_entries = paginator.page(page_number)
        except PageNotAnInteger:
            voucher_entries = paginator.page(1)
        except EmptyPage:
            voucher_entries = paginator.page(paginator.num_pages)

        context = {
            'voucher_detail': voucher_detail,
            'voucher_entries': list(voucher_entries),
            'page': voucher_entries.number,
            'pages': paginator.num_pages,
            'has_previous': voucher_entries.has_previous(),
            'has_next': voucher_entries.has_next()
        }

        if voucher_entries_len == 0:
            context['message'] = 'No se encontraron entradas para este vale.'

        return JsonResponse(context)
    except Voucher.DoesNotExist:
        return JsonResponse({'error': 'Vale no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def export_excel_perpetuo_report(request):
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    lot = request.GET.get('filterLotNumber')

    
    try:
        filters = {
            'voucher__lots__client_diner__client__id': request.GET.get('filterClient'),
            'voucher__lots__client_diner__dining_room__id': request.GET.get('filterDiningRoom'),
            'employee__icontains': request.GET.get('filterEmployeeName'),
            'voucher__status': request.GET.get('filterStatus'),
            'voucher__lots__voucher_type_id': 2,  # Tipo de vale perpetuo
            'voucher__folio__icontains': request.GET.get('filterVoucherFolio'),
            'voucher__lots__id': lot
        }

        # Convertir las fechas a objetos datetime conscientes de la zona horaria
        start_date_str = request.GET.get('filterStartDate')
        end_date_str = request.GET.get('filterEndDate')

        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            start_date = timezone.make_aware(start_date)
        else:
            start_date = None

        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            end_date = timezone.make_aware(end_date)
        else:
            end_date = None

        # Filtro de entradas solo cuando se han proporcionado fechas
        if start_date:
            filters['created_at__gte'] = start_date
        if end_date:
            filters['created_at__lte'] = end_date

        filters = {k: v for k, v in filters.items() if v}

        entry_perpetual = Entry.objects.select_related(
            'voucher__lots__client_diner__client',
            'voucher__lots__client_diner__dining_room'
        ).filter(**filters).values(
            client_company=F('voucher__lots__client_diner__client__company'),
            client_name=F('voucher__lots__client_diner__client__name'),
            client_lastname=F('voucher__lots__client_diner__client__lastname'),
            client_second_lastname=F('voucher__lots__client_diner__client__second_lastname'),
            dining_room_name=F('voucher__lots__client_diner__dining_room__name'),
            employee_name=F('voucher__employee'),
            voucher_folio=F('voucher__folio'),
            voucher_status=F('voucher__status'),
            entry_created_at=F('created_at')
        ).order_by('-created_at')

        # Create a DataFrame from the queryset
        df = pd.DataFrame(list(entry_perpetual))

        # Remove timezone information from datetime objects
        if 'entry_created_at' in df.columns:
            df['entry_created_at'] = df['entry_created_at'].apply(lambda x: format_date(x) if pd.notnull(x) else "Sin fecha")

        # Si el estado del vale es verdadero asignar "Activo" y si es falso asignar "Inactivo"
        df['voucher_status'] = df['voucher_status'].apply(lambda x: 'Activo' if x else 'Inactivo')

        # Define the custom headers for the Excel sheet
        headers = [
            'Compañía Cliente',
            'Nombre Cliente',
            'Apellido Paterno Cliente',
            'Apellido Materno Cliente',
            'Nombre Comedor',
            'Nombre Empleado',
            'Folio Vale',
            'Estado Vale',
            'Fecha de Ingreso'
        ]

        # Create a new Excel workbook and add two sheets
        wb = openpyxl.Workbook()

        # Sheet 1: Perpetual Report
        ws1 = wb.active
        ws1.title = "Perpetual Report"
        ws1.append(headers)
        add_styles(ws1, headers)

        for row in dataframe_to_rows(df, index=False, header=False):
            ws1.append(row)

        # Sheet 2: Perpetual Report Summary (simplified version)
        ws2 = wb.create_sheet(title="Summary Report")
        simplified_headers = [
            'Compañía Cliente',
            'Nombre Cliente',
            'Apellido Paterno Cliente',
            'Apellido Materno Cliente',
            'Nombre Comedor',
            'Folio Vale',
            'Nombre Empleado',
            'Estado Vale',
            'Número de Usos'
        ]
        ws2.append(simplified_headers)
        add_styles(ws2, simplified_headers)

        # Generate a summary report (aggregating by employee)
        filters = {
            'lots__client_diner__client__id': request.GET.get('filterClient'),
            'lots__client_diner__dining_room__id': request.GET.get('filterDiningRoom'),
            'employee__icontains': request.GET.get('filterEmployeeName'),
            'status': request.GET.get('filterStatus'),
            'lots__voucher_type_id': 2,  # Tipo de vale perpetuo
            'folio__icontains': request.GET.get('filterVoucherFolio'),
            'lots__id': lot
        }
                # Convert date filters to timezone-aware datetime
        if filters.get('created_at__gte'):
            filters['created_at__gte'] = timezone.make_aware(datetime.strptime(filters['created_at__gte'], '%Y-%m-%d'))
        if filters.get('created_at__lte'):
            filters['created_at__lte'] = timezone.make_aware(datetime.strptime(filters['created_at__lte'], '%Y-%m-%d'))
        filters = {k: v for k, v in filters.items() if v}

        # Obtener los vales de tipo 1
        perpetual_report_summary = Voucher.objects.select_related(
            'lots__client_diner__client',
            'lots__client_diner__dining_room'
        ).filter(**filters).values(
            client_company=F('lots__client_diner__client__company'),
            client_name=F('lots__client_diner__client__name'),
            client_lastname=F('lots__client_diner__client__lastname'),
            client_second_lastname=F('lots__client_diner__client__second_lastname'),
            dining_room_name=F('lots__client_diner__dining_room__name'),
            voucher_folio=F('folio'),
            employee_name=F('employee'),
            voucher_status=F('status')
        ).annotate(
            entry_count=Count('entry_voucher')
        ).order_by('voucher_folio')

        summary_df = pd.DataFrame(list(perpetual_report_summary))

        # Si el estado del vale es verdadero asignar "Activo" y si es falso asignar "Inactivo"
        summary_df['voucher_status'] = summary_df['voucher_status'].apply(lambda x: 'Habilitado' if x else 'Deshabilitado')

        for row in dataframe_to_rows(summary_df, index=False, header=False):
            ws2.append(row)

        # Create a BytesIO buffer to hold the Excel file
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Set the response content type to Excel
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=perpetual_report.xlsx'

        return response
    except Exception as e:
        return JsonResponse({'error': 'No hay registros'}, status=500)

# ===================== GENERAR VALES UNICOS ===================== #

@csrf_exempt
def generate_unique_voucher(request):
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)
    
    try:
        data = json.loads(request.body)
        client_id = data.get("client_id")
        dining_room_id = data.get("dining_room_id")
        quantity = data.get("quantity")
        voucher_type = data.get("voucher_type")
        

        if not client_id or not dining_room_id or not quantity or not voucher_type:
            return JsonResponse({"error": "client_id, type, dining_room_id y quantity son requeridos"}, status=400)
        
        if type(quantity) != int:
            return JsonResponse({"error": "quantity debe ser un número entero"}, status=400)
        
        if quantity > 9999:
            return JsonResponse({"error": "La cantidad no puede ser mayor a 9999"}, status=400)

        if type(client_id) != int:
            return JsonResponse({"error": "client_id debe ser un número entero"}, status=400)
        
        if type(dining_room_id) != int:
            return JsonResponse({"error": "dining_room_id debe ser un número entero"}, status=400)
        
        if voucher_type not in ["UNICO", "PERPETUO"]:
            return JsonResponse({"error": "El tipo de vale debe ser 'UNICO' o 'PERPETUO'"}, status=400)
    
        
        voucher_type_obj = VoucherType.objects.filter(description=voucher_type).first()     
        client_dinner = ClientDiner.objects.filter(client_id=client_id, dining_room_id=dining_room_id).first()

        if not client_dinner:
            return JsonResponse({"error": "No se encontró la relación entre cliente y comedor"}, status=400)
        
        if not client_dinner.client.status:
            return JsonResponse({"error": f"El cliente {client_dinner.client.company} ha sido desactivado."}, status=400)
        
        if not client_dinner.dining_room.status:
            return JsonResponse({"error": f"El comedor {client_dinner.dining_room.name} ha sido desactivado."}, status=400)
    
        if not client_dinner.status:
            return JsonResponse({"error": f"El uso del comedor por parte de  {client_dinner.client.company} ha sido desactivado."}, status=400)
        
        diningroom = client_dinner.dining_room
        
        clean_pdf_dir()
        
        with transaction.atomic():
            lots = Lots(
                client_diner=client_dinner,
                voucher_type=voucher_type_obj,
                quantity=quantity,
                created_by=request.user
            )
            lots.save()
            
            vouchers: list[Voucher] = []
            
            for _ in range(quantity):
                voucher = Voucher(lots=lots)
                voucher.save()
                vouchers.append(voucher)
            
            qr_paths = prepare_qrs(vouchers, lots.id, diningroom.name)
            
            
            filename = f'/LOT-{lots.id}.pdf'
            generate_qrs_pdf(qr_paths, filename, lots.voucher_type.description)
            
            context = {
                "lot_id": lots.id,
                "pdf": filename,
                "email": client_dinner.client.email,
                "message": "Vales generados con éxito"
            }
            
            for qr in qr_paths:
                os.remove(qr[0])


            return JsonResponse(context)

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

#Funcion para devolver la ruta al pdf del lote
@csrf_exempt
def get_lot_pdf(request):
    lot_id = request.GET.get('lot_id')
    
    if not lot_id:
        return JsonResponse({"error": "El id del lote es requerido"}, status=400)
    
    lot = Lots.objects.filter(id=lot_id).first()

    if not lot:
        return JsonResponse({"error": "El lote especificado no existe"}, status=404)
    
    #Verifica si el pdf de dicho lote existe
    filepath = verify_lot_pdf_exists(lot.id)
    if filepath and lot.voucher_type.description == "UNICO":
        url = prepare_url_pdf(filepath)
        email = lot.email if lot.email else lot.client_diner.client.email
    
        return JsonResponse({"pdf": url, "email": email})
    
    vouchers = Voucher.objects.filter(lots=lot)
    
    if not vouchers:
        return JsonResponse({"error": "No se encontraron vales para el lote especificado"}, status=404)
    
    qr_paths = prepare_qrs(vouchers, lot.id, lot.client_diner.dining_room.name)
    
    filename = create_lot_pdf_name(lot.id)
    filepath = generate_qrs_pdf(qr_paths, filename, lot.voucher_type.description)
    
    url = prepare_url_pdf(filepath)

    email = lot.email if lot.email else lot.client_diner.client.email
    
    return JsonResponse({"pdf": url, "email": email})
    
    
@csrf_exempt
def generate_perpetual_voucher(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        data = json.loads(request.body)
        client_id = data.get('client_id')
        dining_room_id = data.get('dining_room_id')
        quantity = data.get('quantity')
        
        
        if not client_id or not dining_room_id or not quantity :
            return JsonResponse({'error': 'client_id, dining_room_id, quantity y employees son requeridos'}, status=400)
        
        if type(quantity) != int:
            return JsonResponse({'error': 'La cantidad debe ser un número entero'}, status=400)
        
        if type(client_id) != int:
            return JsonResponse({'error': 'El id del cliente debe ser un número entero'}, status=400)

        if type(dining_room_id) != int:
            return JsonResponse({'error': 'El id del comedor debe ser un número entero'}, status=400)  
        
        
        if quantity > 99:
            return JsonResponse({'error': 'La cantidad no puede ser mayor a 99'}, status=400)

        
        perpetual_voucher = VoucherType.objects.filter(description="PERPETUO").first()
        
        client_dinner = ClientDiner.objects.filter(client_id=client_id, dining_room_id=dining_room_id).first()

        if not client_dinner:
            return JsonResponse({"error": "No se encontró la relación entre cliente y comedor"}, status=400)
        
        if not client_dinner.client.status:
            return JsonResponse({"error": f"El cliente {client_dinner.client.company} ha sido desactivado."}, status=400)

        if not client_dinner.dining_room.status:
            return JsonResponse({"error": f"El comedor {client_dinner.dining_room.name} ha sido desactivado."}, status=400)
        
        if not client_dinner.status:
            return JsonResponse({"error": f"El uso del comedor por parte de  {client_dinner.client.company} ha sido desactivado."}, status=400)
        


        with transaction.atomic():
            lot = Lots(
                client_diner=client_dinner,
                voucher_type=perpetual_voucher,
                quantity=quantity,
                created_by=request.user
            )
            lot.save()

            vouchers = []
            
            for _ in range(quantity):
                voucher = Voucher(lots=lot)
                voucher.save()
                vouchers.append(voucher)

            

            email = lot.email if lot.email else lot.client_diner.client.email
            
            
            vouchers_objects = [{"id": voucher.id, "folio": voucher.folio} for voucher in vouchers]

        return JsonResponse({'message': 'Vales generados con éxito', "lot": lot.id, "vouchers": vouchers_objects, "email": email})
    except Exception as err:
        return JsonResponse({"error": str(err)}, status=500)
        
@csrf_exempt
def send_lot_file_email(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        data = json.loads(request.body)
        email = data.get('email')
        lot = data.get('lot_id')
        
        
        #validate email being not null
        if not email:
            return JsonResponse({'error': 'El email es requerido'}, status=400)
        
        #validate lots being not null
        if not lot:
            return JsonResponse({'error': 'El lote es requerido'}, status=400)
        
        #validate email being a string
        if type(email) != str:
            return JsonResponse({'error': 'El email debe ser una cadena de texto'}, status=400)
        
        #validate lot being an int
        if type(lot) != int:
            return JsonResponse({'error': 'El lote debe ser un número entero'}, status=400)
        
        #Validate email being less or equal to 100
        if len(email) > 100:
            return JsonResponse({'error': 'El email no puede ser mayor a 100 caracteres'}, status=400)
        
        lot_object = Lots.objects.get(id=lot)
        
        if not lot_object:
            return JsonResponse({'error': 'El lote no existe'}, status=404)
        
        sender_email = config('EMAIL') 
        sender_password = config('EMAIL_PASSWORD')

        if not sender_email or not sender_password:
            return JsonResponse({"error": "No se tiene configurado el email para enviar correos"},status=500)
        

        email_message = EmailMessage()

        filepath = None
        try:
            filepath = generate_lot_pdf(lot, lot_object.voucher_type.description == "UNICO") 
        except Exception as err:
            return JsonResponse({"error": "Hubo un error generando el pdf"}, status=500)


        with open(filepath, "rb") as f:
            pdf_data = f.read()
            email_message.add_attachment(pdf_data, maintype="application", subtype="pdf", filename=f"LOT-{lot}.pdf")
        
        email_context = ssl.create_default_context()
        subject = f'Archivo de lote {lot}'
        email_message['From'] = sender_email
        email_message['To'] = email
        email_message['Subject'] = subject

        
        with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=email_context) as server:
            try:
                server.login(sender_email, sender_password)
            except Exception as err:
                return JsonResponse({"error": "No se puedo iniciar sesión en el servidor de correos"}, status=500)

            try:
                server.sendmail(sender_email, email, email_message.as_string())
            except Exception as err:
                return JsonResponse({"error": "No se pudo enviar el correo"})
         
        lot_object.email = email
        lot_object.save()
        return JsonResponse({"message": "Email enviado con éxito"})
        
    
    except Exception as err:
        return JsonResponse({'error': str(err)}, status=500)

@csrf_exempt
def generate_perpetual_voucher_qr(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        data = json.loads(request.body)
        voucher_id = data.get('voucher_id')
        
        voucher = Voucher.objects.filter(id=voucher_id).first()
        
        if not voucher:
            return JsonResponse({'error': 'El vale no existe'}, status=404)
        
        perpetual_type = VoucherType.objects.filter(description="PERPETUO").first()
        
        if not perpetual_type:
            return JsonResponse({'error': 'El tipo de vale perpetuo no existe'}, status=404)

        if voucher.lots.voucher_type.id != perpetual_type.id:
            return JsonResponse({'error': 'El vale no es de tipo perpetuo'}, status=400)

        qr_path = prepare_qr(voucher)
    
        filepath = None
        try: 
            filepath = generate_perpetual_voucher_pdf(voucher, qr_path)
        except:
            return JsonResponse({'error': 'Error al generar el PDF'}, status=500)
        
        match = re.search(r"\\static\\.*", filepath)
        relative_path = match.group(0)[1:] if match else filepath
        relative_path = relative_path.replace("\\", "/")  


        return JsonResponse({'filepath': relative_path, 'message': 'QR generado con éxito'})

    except:
        return JsonResponse({'error': 'Error al generar el QR'}, status=500)
        
@csrf_exempt
def change_voucher_employee(request):      
    if request.method != 'PUT':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    try:
        data = json.loads(request.body)
        voucher_id = data.get('voucher_id')
        employee = data.get('employee')
        
        if not voucher_id or not employee:        
            return JsonResponse({'error': 'voucher_id y employee son requeridos'}, status=400)

        voucher = Voucher.objects.filter(id=voucher_id).first()
        
        if not voucher:
            return JsonResponse({'error': 'El vale no existe'}, status=404)
        
        perpetual_voucher_type = VoucherType.objects.filter(description="PERPETUO").first()
        
        if voucher.lots.voucher_type.id != perpetual_voucher_type.id:
            return JsonResponse({'error': 'El vale no es de tipo perpetuo'}, status=400)

        voucher.employee = employee
        voucher.save()

        
        return JsonResponse({"message": "Nombre del vale actualizado con éxito" })
    except Exception as err:
        return JsonResponse({'error': str(err)}, status=500)

        
      
    

# ===================== ENTRADAS ===================== #
@csrf_exempt
def entradas_view(request):
    try:
        user = request.user.id        
        dining_room = DiningRoom.objects.filter(in_charge_id=user, status=True).first()        

        if dining_room:
            client_diner_dining_room = ClientDiner.objects.filter(dining_room=dining_room).first()            

            if client_diner_dining_room:
                response_data = {
                    'has_dining_room': True,
                    'dining_room': {
                        'name': dining_room.name,
                        'client_company': client_diner_dining_room.client.company
                    }
                }
            else:
                response_data = {
                    'has_dining_room': False
                }
        else:
            response_data = {
                'has_dining_room': False
            }

        return JsonResponse(response_data)
    except Exception as e:        
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def entradas_te_view(request):
    try:
        user = request.user.id        
        dining_room = DiningRoom.objects.filter(in_charge_id=user, status=True).first()        

        if dining_room:
            client_diner_dining_room = ClientDiner.objects.filter(dining_room=dining_room).first()            

            if client_diner_dining_room:
                response_data = {
                    'has_dining_room': True,
                    'dining_room': {
                        'name': dining_room.name,
                        'client_company': client_diner_dining_room.client.company
                    }
                }
            else:
                response_data = {
                    'has_dining_room': False
                }
        else:
            response_data = {
                'has_dining_room': False
            }

        return JsonResponse(response_data)
    except Exception as e:        
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def get_informacion_comedor_entradas(request):
    try:
        user_id = request.user.id
        dining_rooms = DiningRoom.objects.filter(
            status=True,
            in_charge__status=True,
            in_charge_id=user_id
        ).select_related('in_charge', 'client_diner_dining_room__client').values(
            'name',
            'client_diner_dining_room__client__company'
        )

        result = [
            {
                'dining_room_name': dr['name'],
                'client_company': dr['client_diner_dining_room__client__company']
            }
            for dr in dining_rooms
        ]

        return JsonResponse({'data': result})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def manejar_vale_unico(voucher):
    # Verificar si el vale esta activo
    if not voucher.status:
        return JsonResponse({"message": "Este vale ya fue usado", "status": "info"}, status=400)

    # Cambiar el estatus del vale a utilizado
    voucher.status = False
    voucher.save()

    # Definir la zona horaria de Arizona
    arizona_tz = pytz.timezone('America/Phoenix')

    # Guardar la información de la entrada con la zona horaria de Arizona
    entry = Entry(
        voucher=voucher,
        client_diner=voucher.lots.client_diner,
        created_at=timezone.now().astimezone(arizona_tz)
    )
    entry.save()

    return JsonResponse({"message": "Bienvenido al comedor", "status": "success"})


def manejar_vale_perpetuo(voucher):
    # Verificar si el vale está activo
    if not voucher.status:
        return JsonResponse({"message": "Este vale está inactivo", "status": "info"}, status=400)

    # Definir la zona horaria de Arizona
    arizona_tz = pytz.timezone('America/Phoenix')
    now = timezone.now().astimezone(arizona_tz)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = now.replace(hour=23, minute=59, second=59, microsecond=999999)

    # Verificar si el vale ya ha sido utilizado hoy
    if Entry.objects.filter(voucher=voucher, created_at__range=(today_start, today_end)).exists():
        return JsonResponse({"message": "Este vale ya ha sido utilizado hoy", "status": "info"}, status=400)

    # Guardar la información de la entrada con la zona horaria de Arizona
    entry = Entry(
        voucher=voucher,
        client_diner=voucher.lots.client_diner,
        created_at=now
    )
    entry.save()

    return JsonResponse({"message": "Bienvenido al comedor", "status": "success"})


@csrf_exempt
def validar_vale(request):
    try:        
        data = json.loads(request.body)
        folio = data.get("folio")

        # Verificar si el folio fue proporcionado
        if not folio:
            return JsonResponse({"message": "Debes ingresar un folio", "status": "danger"}, status=400)

        # Verificar que el folio siga la estructura Lote-número (por ejemplo, "2-7")
        folio_regex = re.compile(r'^\d+-\d+$')
        if not folio_regex.match(folio):
            return JsonResponse({"message": "Formato incorrecto", "status": "info"}, status=400)

        voucher = Voucher.objects.filter(folio=folio).select_related('lots__voucher_type', 'lots__client_diner__dining_room').first()

        # Verificar que el vale exista
        if not voucher:
            return JsonResponse({"message": "Vale no encontrado", "status": "danger"}, status=404)

        # Verificar si el usuario tiene un comedor asignado
        user_id = request.user.id
        dining_room = DiningRoom.objects.filter(in_charge_id=user_id).first()
        if not dining_room:
            return JsonResponse({'message': 'No tienes un comedor asignado', "status": "danger"}, status=403)
        
        # Verificar si el comedor asignado está activo
        if not dining_room.status:
            return JsonResponse({'message': 'El comedor asignado está inactivo', "status": "danger"}, status=403)

        # Verificar si el vale corresponde al comedor asignado
        if voucher.lots.client_diner.dining_room != dining_room:
            return JsonResponse({"message": "El vale no corresponde al comedor asignado", "status": "danger"}, status=403)

        # Identificar el tipo de vale y manejar la lógica correspondiente
        if voucher.lots.voucher_type.description == "UNICO":
            return manejar_vale_unico(voucher)
        elif voucher.lots.voucher_type.description == "PERPETUO":
            return manejar_vale_perpetuo(voucher)
        else:
            return JsonResponse({"message": "Tipo de vale desconocido", "status": "danger"}, status=400)

    except json.JSONDecodeError:
        return JsonResponse({"error": "El cuerpo de la solicitud debe ser un JSON válido"}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def validar_empleado(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            employeed_code = data.get('employeed_code')

            # Verificar si el código de empleado fue proporcionado
            if not employeed_code:
                return JsonResponse({'message': 'El código de empleado es requerido', "status": "danger"}, status=400)

            # Verificar si el empleado existe
            employee = Employee.objects.filter(employeed_code=employeed_code).first()
            if not employee:
                return JsonResponse({'message': 'Empleado no encontrado', "status": "danger"}, status=404)

            # Verificar si el empleado está activo
            if not employee.status:
                return JsonResponse({'message': 'Empleado inactivo', "status": "danger"}, status=400)
 
            # Verificar si el usuario tiene un comedor asignado
            user_id = request.user.id
            dining_room = DiningRoom.objects.filter(in_charge_id=user_id).first()
            if not dining_room:
                return JsonResponse({'message': 'No tienes un comedor asignado', "status": "danger"}, status=403)
            
            # Verificar si el comedor asignado está activo
            if not dining_room.status:
                return JsonResponse({'message': 'El comedor asignado está inactivo', "status": "danger"}, status=403)

            # Verificar si el empleado tiene acceso al comedor asignado
            employee_client_diner = EmployeeClientDiner.objects.filter(employee=employee, client_diner__dining_room=dining_room).first()
            if not employee_client_diner:
                return JsonResponse({'message': 'El empleado no tiene acceso a este comedor', 'status': "danger"}, status=403)

            # Definir la zona horaria de Arizona
            arizona_tz = pytz.timezone('America/Phoenix')
            now = timezone.now().astimezone(arizona_tz)
            today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
            today_end = now.replace(hour=23, minute=59, second=59, microsecond=999999)

            # Verificar si el empleado ya ha registrado una entrada hoy
            if Entry.objects.filter(employee_client_diner__employee=employee, created_at__range=(today_start, today_end)).exists():
                return JsonResponse({'message': 'El empleado ya ha registrado una entrada hoy', "status": "info"}, status=400)

            # Registrar la entrada
            entry = Entry(
                employee_client_diner=employee_client_diner,
                client_diner=employee_client_diner.client_diner,
                created_at=now
            )
            entry.save()

            return JsonResponse({'message': 'Bienvenido al comedor', 'status': 'success'}, status=200)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'El cuerpo de la solicitud debe ser un JSON válido'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def validar_empleado_te(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            employeed_code = data.get('employeed_code')

            # Verificar si el código de empleado fue proporcionado
            if not employeed_code:
                return JsonResponse({'message': 'El código de empleado es requerido', "status": "danger"}, status=400)

            # Verificar si el empleado existe
            employee = Employee.objects.filter(employeed_code=employeed_code).first()
            if not employee:
                return JsonResponse({'message': 'Empleado no encontrado', "status": "danger"}, status=404)

            # Verificar si el empleado está activo
            if not employee.status:
                return JsonResponse({'message': 'Empleado inactivo', "status": "danger"}, status=400)
 
            # Verificar si el usuario tiene un comedor asignado
            user_id = request.user.id
            dining_room = DiningRoom.objects.filter(in_charge_id=user_id).first()
            if not dining_room:
                return JsonResponse({'message': 'No tienes un comedor asignado', "status": "danger"}, status=403)
            
            # Verificar si el comedor asignado está activo
            if not dining_room.status:
                return JsonResponse({'message': 'El comedor asignado está inactivo', "status": "danger"}, status=403)

            # Verificar si el empleado tiene acceso al comedor asignado
            employee_client_diner = EmployeeClientDiner.objects.filter(employee=employee, client_diner__dining_room=dining_room).first()
            if not employee_client_diner:
                return JsonResponse({'message': 'El empleado no tiene acceso a este comedor', 'status': "danger"}, status=403)

            # Definir la zona horaria de Arizona
            arizona_tz = pytz.timezone('America/Phoenix')
            now = timezone.now().astimezone(arizona_tz)
            today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
            today_end = now.replace(hour=23, minute=59, second=59, microsecond=999999)

            # Registrar la entrada
            entry = Entry(
                employee_client_diner=employee_client_diner,
                client_diner=employee_client_diner.client_diner,
                created_at=now
            )
            entry.save()

            return JsonResponse({'message': 'Bienvenido al comedor', 'status': 'success'}, status=200)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'El cuerpo de la solicitud debe ser un JSON válido'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def get_last_entries(request):
    if request.method != 'GET':
        return JsonResponse({'error': 'Metodo no permitido'}, status=405)
    
    try:
        last_entries = request.GET.get('last_entries')
        dining_room = request
        
        if not last_entries:
            return JsonResponse({'error': 'last_entries es requerido'}, status=400)
        
        last_entries = int(last_entries)
        
        if last_entries > 100:
            return JsonResponse({'error': 'last_entries no puede ser mayor a 100'}, status=400)
        
        user = CustomUser.objects.filter(id=request.user.id).first()

                
        if not user:
            return JsonResponse({'error': 'El usuario no existe'}, status=404)

        dining_room = DiningRoom.objects.filter(in_charge=user).first()

        
        if not dining_room:
            return JsonResponse({'error': 'El usuario no tiene un comedor asignado'}, status=404)
        
        client_diner = ClientDiner.objects.filter(dining_room=dining_room).first()

        if not client_diner:
            return JsonResponse({'error': 'El comedor no tiene clientes asignados'}, status=404)
        
        
        
        entries = Entry.objects.filter(client_diner=client_diner).order_by('-created_at')[:last_entries]

        arizona_tz = pytz.timezone('America/Phoenix')
        entries_json = [
            {
            "employee": entry.employee_client_diner.employee.name if entry.employee_client_diner else None,
            "datetime": entry.created_at.astimezone(arizona_tz).strftime('%Y-%m-%d %H:%M:%S'),
            "voucher": entry.voucher.folio if entry.voucher else None,
            "voucher_type": entry.voucher.lots.voucher_type.description if entry.voucher else None
            } for entry in entries
        ]

        return JsonResponse({'entries': entries_json})
    
    except Exception as err:
        print(str(err))
        return JsonResponse({'error': 'Error al obtener las entradas'}, status=500)
        

    
# ===================== ADMINISTRADOR DE VALES ===================== #
@csrf_exempt
def get_voucher_lots(request):
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    try:
        # Filtros
        lot_id = request.GET.get('lot_id')
        voucher_type = request.GET.get('voucher_type')
        page_number = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 10))

        lots_query = Lots.objects.select_related('voucher_type', 'created_by', 'client_diner__client', 'client_diner__dining_room').values(
            'id',
            'voucher_type__description',
            'quantity',
            'email',
            'created_at',
            'created_by__username',
            'client_diner__client__company',
            'client_diner__client__email',
            'client_diner__dining_room__name'
        ).order_by('-id')

        if lot_id:
            lots_query = lots_query.filter(id__icontains=lot_id)
        if voucher_type:
            lots_query = lots_query.filter(voucher_type__description__iexact=voucher_type)

        lots = [
            {
                'id': lot['id'],
                'voucher_type': lot['voucher_type__description'] or 'N/A',
                'quantity': lot['quantity'] or 'N/A',
                'email': lot['email'] or lot['client_diner__client__email'] or 'N/A',
                'created_at': lot['created_at'].strftime('%Y-%m-%d %H:%M:%S') if lot['created_at'] else 'N/A',
                'created_by': lot['created_by__username'] or 'N/A',
                'client': lot['client_diner__client__company'] or 'N/A',
                'dining_room': lot['client_diner__dining_room__name'] or 'N/A'
            }
            for lot in lots_query
        ]

        paginator = Paginator(lots, page_size)
        page_obj = paginator.get_page(page_number)

        return JsonResponse({
            'lots': list(page_obj),
            'page': page_obj.number,
            'pages': paginator.num_pages,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous()
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def get_vouchers_by_lot(request):
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    try:
        lot_id = request.GET.get('lot_id')
        folio = request.GET.get('folio')
        page_number = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 10))

        if not lot_id:
            return JsonResponse({'error': 'El parámetro lot_id es obligatorio'}, status=400)
        
        # Filtros dinámicos
        filters = {'lots_id': lot_id}

        if folio:  # Solo agregar el filtro de folio si se proporciona
            filters['folio__icontains'] = folio

        vouchers_query = Voucher.objects.filter(**filters).select_related(
            'lots__voucher_type',
            'lots__client_diner__client'
        ).values(
            'id',
            'folio',
            'employee',
            'status',
            'lots__voucher_type__description',
            'lots__email',
            'lots__client_diner__client__company',
            'lots__client_diner__client__email',
            'lots__client_diner__dining_room__name'
        ).order_by('-id')

        vouchers = [
            {
                'id': voucher['id'],
                'folio': voucher['folio'] or 'N/A',
                'employee': voucher['employee'] or '',
                'status': 1 if voucher['status'] else 0,
                'voucher_type': voucher['lots__voucher_type__description'] or 'N/A',
                'client': voucher['lots__client_diner__client__company'] or 'N/A',
                'email': voucher['lots__email'] or voucher['lots__client_diner__client__email'] or 'N/A',
                'dining_room': voucher['lots__client_diner__dining_room__name'] or 'N/A'
            }
            for voucher in vouchers_query
        ]

        paginator = Paginator(vouchers, page_size)
        page_obj = paginator.get_page(page_number)

        return JsonResponse({
            'vouchers': list(page_obj),
            'page': page_obj.number,
            'pages': paginator.num_pages,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous()
        })
    except Exception as e:
        print(e)
        return JsonResponse({'error': str(e)}, status=500)
    
@csrf_exempt
def search_pdf_qr_perpetual_voucher_and_generate(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        data = json.loads(request.body)
        voucher_folio = data.get('voucher_folio')
        
        if not voucher_folio:
            return JsonResponse({'error': 'voucher_folio es requerido'}, status=400)
        
        voucher = Voucher.objects.filter(folio=voucher_folio).first()

        if not voucher:
            return JsonResponse({'error': 'Vale no encontrado'}, status=404)

        if voucher.lots.voucher_type.description != 'PERPETUO':
            return JsonResponse({'error': 'El vale no es de tipo perpetuo'}, status=400)
        
        existing_filepath = verify_voucher_pdf_exists(voucher.id)
        if existing_filepath:
            filename = existing_filepath.split('/')[-1]
            return JsonResponse({'pdf': f'/static/pdfs/{filename}', 'message': 'PDF ya existente'})
        
        qr_path = prepare_qr(voucher)
    
        filepath = None
        try: 
            filepath = generate_perpetual_voucher_pdf(voucher, qr_path)
        except:
            return JsonResponse({'error': 'Error al generar el PDF'}, status=500)

        match = re.search(r"\\static\\.*", filepath)
        relative_path = match.group(0)[1:] if match else filepath
        relative_path = relative_path.replace("\\", "/")  


        return JsonResponse({'pdf': relative_path, 'message': 'PDF generado con éxito'})
    except Exception as e:
        print(e)
        return JsonResponse({'error': str(e)}, status=500)
    
@csrf_exempt
def change_voucher_status(request):      
    if request.method != 'PUT':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    try:
        data = json.loads(request.body)
        voucher_id = data.get('voucher_id')
        status = data.get('status')
        
        if not voucher_id or not isinstance(status, bool):
            return JsonResponse({'error': 'voucher_id y status booleano son requeridos'}, status=400)

        voucher = Voucher.objects.filter(id=voucher_id).first()
        
        if not voucher:
            return JsonResponse({'error': 'El vale no existe'}, status=404)
        
        perpetual_voucher_type = VoucherType.objects.filter(description="PERPETUO").first()
        
        if voucher.lots.voucher_type.id != perpetual_voucher_type.id:
            return JsonResponse({'error': 'El vale no es de tipo perpetuo'}, status=400)

        voucher.status = status 
        voucher.save()

        
        return JsonResponse({"message": "Nombre del vale actualizado con éxito" })
    except Exception as err:
        return JsonResponse({'error': str(err)}, status=500)